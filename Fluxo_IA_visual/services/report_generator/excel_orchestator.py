# Fluxo_IA_visual/services/report_generator/excel_orchestrator.py
import io
import logging
from openpyxl import Workbook

from .sheets import s01_executive_summary
from .sheets import s02_billing_history
from .sheets import s03_forecast
from .sheets import s04_raw_data
from .sheets import s05_financial_stmts
from .sheets import s06_buro_detail
from .sheets import s07_business_network
from .sheets import s08_products_services

logger = logging.getLogger(__name__)

class ExcelReportBuilder:
    """
    Orquestador principal para la generación del reporte Excel.
    Implementa el patrón Builder delegando cada hoja a su propio módulo.
    """
    def __init__(self, data: dict):
        self.data = data
        self.wb = Workbook()

    def build(self) -> bytes:
        logger.info(f"[{self.data.get('rfc', 'N/A')}] Iniciando construcción de reporte Excel...")
        
        # 1. Crear hojas base
        ws1 = self.wb.active
        ws1.title = "Resumen Ejecutivo"
        
        ws2 = self.wb.create_sheet("Facturación en el tiempo")
        ws3 = self.wb.create_sheet("Proyecciones (Escenarios)")
        ws4 = self.wb.create_sheet("Raw Data")
        ws5 = self.wb.create_sheet("Estados Financieros")
        ws6 = self.wb.create_sheet("Detalle de Buró")
        ws7 = self.wb.create_sheet("Red de Negocios")
        ws8 = self.wb.create_sheet("Productos y Servicios")

        # 2. Ejecutar los builders de cada hoja pasándoles su Worksheet y la Data
        # Descomentaremos estas líneas en las siguientes fases
        
        s01_executive_summary.build(ws1, self.data)
        s02_billing_history.build(ws2, self.data)
        s03_forecast.build(ws3, self.data)
        s04_raw_data.build(ws4, self.data)
        s05_financial_stmts.build(ws5, self.data)
        s06_buro_detail.build(ws6, self.data)
        s07_business_network.build(ws7, self.data)
        s08_products_services.build(ws8, self.data)

        # 3. Guardar en memoria y retornar bytes
        output = io.BytesIO()
        self.wb.save(output)
        
        logger.info("Excel ensamblado correctamente en memoria.")
        return output.getvalue()