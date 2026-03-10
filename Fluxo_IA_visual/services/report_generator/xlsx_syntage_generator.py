import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle, Border, Side
from datetime import datetime

def generar_excel_syntage(data: dict) -> bytes:
    wb = Workbook()
    
    # --- ESTILOS ---
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    sub_header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    currency_style = NamedStyle(name='currency_style', number_format='$#,##0.00')
    percent_style = NamedStyle(name='percent_style', number_format='0.00%')
    date_style = NamedStyle(name='date_style', number_format='YYYY-MM-DD')
    alert_fill = PatternFill(start_color="C0504D", end_color="C0504D", fill_type="solid")

    def estilo_header(ws, row_idx, col_start=1, col_end=None):
        if col_end is None: col_end = ws.max_column
        for col in range(col_start, col_end + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

    # ==========================================
    # HOJA 1: RESUMEN EJECUTIVO (MODIFICADA)
    # ==========================================
    ws1 = wb.active
    ws1.title = "Resumen Ejecutivo"
    
    # 1. Datos Contribuyente (+ Antigüedad)
    ws1.append(["DATOS DEL CONTRIBUYENTE"])
    ws1.merge_cells('A1:D1') # Expandimos a 4 columnas por el requerimiento de dividir celdas
    estilo_header(ws1, 1, 1, 4)
    
    # Cálculo Antigüedad
    antiguedad_str = "N/A"
    reg_date_str = data.get("tax_registration_date")
    if reg_date_str:
        try:
            reg_dt = datetime.fromisoformat(reg_date_str.replace("Z", ""))
            years = (datetime.now() - reg_dt).days // 365
            antiguedad_str = f"{years} Años ({reg_date_str[:10]})"
        except: pass

    ws1.append(["Razón Social", data.get("business_name", "N/A"), "Antigüedad SAT", antiguedad_str])
    ws1.append(["RFC", data.get("rfc", "N/A"), "", ""]) # Merge visual opcional

    # 2. Credenciales
    ws1.append([])
    ws1.append(["ESTATUS DE CREDENCIALES"])
    ws1.merge_cells('A5:D5')
    estilo_header(ws1, 5, 1, 4)
    
    ciec = data.get("ciec_info", {})
    buro = data.get("buro_info", {})
    opinion = data.get("compliance_opinion", {})

    # Formato solicitado: Col A: Nombre, Col B: Estatus, Col C: Label Extra, Col D: Valor Extra
    ws1.append(["Credencial", "Estatus", "Detalle / Fecha", "Valor / Score"])
    estilo_header(ws1, 6, 1, 4)

    def format_ciec_date(date_str):
        if not date_str: return ""
        try:
            d = datetime.fromisoformat(str(date_str).replace("Z", "").split(".")[0])
            return f"{d.month}/{d.day}/{d.year}"
        except:
            return date_str
    
    # Fila CIEC
    ws1.append(["CIEC", ciec.get("status"), "Última Rev:", ciec.get("last_check_date")])
    # Fila Ultima extracción
    ws1.append(["CIEC última extracción", ciec.get("status"), "última extracción", format_ciec_date(ciec.get("last_extraction_date"))])
    # Fila Opinión
    ws1.append(["Opinión Cumpl.", opinion.get("status"), "Última Rev:", opinion.get("last_check_date")])
    # Fila Buró (Split Solicitado)
    ws1.append(["Buró de Crédito", buro.get("status"), "Score:", buro.get("score")])

    # 3. Riesgos
    ws1.append([])
    ws1.append(["INDICADORES DE RIESGO"])
    
    # Hacemos que la mezcla de celdas y el estilo sean 100% dinámicos
    fila_actual = ws1.max_row
    ws1.merge_cells(f'A{fila_actual}:D{fila_actual}')
    estilo_header(ws1, fila_actual, 1, 4)
    
    ws1.append(["Indicador", "Valor", "Riesgoso", ""])
    
    risks = data.get("risk_indicators", [])
    if risks and isinstance(risks, list):
        risk_obj = risks[0]
        for k, v in risk_obj.items():
            if isinstance(v, dict):
                es_riesgoso = "SÍ" if v.get("risky") else "NO"
                ws1.append([k, str(v.get("value")), es_riesgoso, ""])
                if v.get("risky"):
                    ws1.cell(row=ws1.max_row, column=3).fill = alert_fill

    # 4. Actividades Económicas
    ws1.append([])
    ws1.append(["ACTIVIDADES ECONÓMICAS"])
    ws1.merge_cells(f'A{ws1.max_row}:D{ws1.max_row}')
    estilo_header(ws1, ws1.max_row, 1, 4)
    
    ws1.append(["Actividad", "Porcentaje", "Fecha Inicio", ""])
    acts = data.get("economic_activities", [])
    for act in acts:
        # Formatear porcentaje
        pct = act.get("percentage", 0)
        pct_str = f"{pct}%"
        ws1.append([act.get("name"), pct_str, act.get("start_date"), ""])

    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 20
    ws1.column_dimensions['C'].width = 20
    ws1.column_dimensions['D'].width = 20

    # 5. Resumen Financiero de Buró 
    summary = buro.get("summary_metrics")
    if summary:
        ws1.append([])
        ws1.append(["RESUMEN DE BURÓ"])
        ws1.merge_cells(f'A{ws1.max_row}:D{ws1.max_row}')
        estilo_header(ws1, ws1.max_row, 1, 4)

        # A. Tendencia de Consultas (Imprime solo el escenario real)
        trend_pct = summary.get("inquiries_trend_pct", 0.0)
        
        if trend_pct > 0:
            ws1.append(["", "Han ido aumentando sus consultas", trend_pct, ""])
            ws1.cell(row=ws1.max_row, column=3).style = percent_style
        elif trend_pct < 0:
            ws1.append(["", "Han ido disminuyendo sus consultas", trend_pct, ""])
            ws1.cell(row=ws1.max_row, column=3).style = percent_style
        else:
            ws1.append(["", "Estables", "0%", ""])
            
        ws1.append([])
        
        # B. Métricas Globales
        metricas_globales = [
            ("Monto Máximo Abierto", summary.get("total_open_max_amount", 0.0), currency_style),
            ("Saldo Vigente", summary.get("total_current_balance", 0.0), currency_style),
            ("Saldo Vencido Total", summary.get("total_past_due", 0.0), currency_style),
            ("Pago Mensual 1", summary.get("monthly_payment_1", 0.0), currency_style),
            ("Pago Mensual 2", summary.get("monthly_payment_2", 0.0), currency_style),
            ("Plazo Ponderado (años)", summary.get("weighted_term_years", 0.0), '0.00')
        ]
        
        for titulo, valor, estilo in metricas_globales:
            ws1.append([titulo, valor, "", ""])
            celda_valor = ws1.cell(row=ws1.max_row, column=2)
            if isinstance(estilo, str): 
                celda_valor.number_format = estilo
            else:
                celda_valor.style = estilo

        ws1.append([])
        
        # C. Cubetas de Morosidad (Buckets)
        def print_bucket(label, key):
            b_data = summary.get(key, {})
            amt = b_data.get("amount", 0.0)
            pct = b_data.get("percentage", 0.0)
            
            # Si es cero, ponemos un guión para que se vea limpio
            val_str = amt if amt > 0 else "-"
            ws1.append([label, val_str, pct if amt > 0 else "", ""])
            
            if amt > 0:
                ws1.cell(row=ws1.max_row, column=2).style = currency_style
                ws1.cell(row=ws1.max_row, column=3).style = percent_style
            else:
                ws1.cell(row=ws1.max_row, column=2).alignment = Alignment(horizontal='right')

        print_bucket("saldoVencidoDe1a29Dias", "bucket_1_29")
        print_bucket("saldoVencidoDe30a59Dias", "bucket_30_59")
        print_bucket("saldoVencidoDe60a89Dias", "bucket_60_89")
        print_bucket("saldoVencidoDe90a119Dias", "bucket_90_119")
        print_bucket("saldoVencidoDe120a179Dias", "bucket_120_179")
        print_bucket("saldoVencidoDe180DiasOMas", "bucket_180_plus")
    
    # 6. Resumen Financiero (vista tipo syntage)
    financials = data.get("financial_ratios_history", [])
    if financials:
        # Ordenamos años de menor a mayor (ej. 2023, 2024)
        financials_sorted = sorted(financials, key=lambda x: str(x.get("year", "")))
        years = [str(f.get("year", "N/A")) for f in financials_sorted]
        max_cols = len(years) + 1

        ws1.append([])
        ws1.append(["RESUMEN FINANCIERO (BALANCE Y RESULTADOS)"] + [""] * (len(years) - 1))
        ws1.merge_cells(start_row=ws1.max_row, start_column=1, end_row=ws1.max_row, end_column=max_cols)
        estilo_header(ws1, ws1.max_row, 1, max_cols)

        # Encabezados de tabla (Concepto, 2023, 2024...)
        ws1.append(["Concepto"] + years)
        fila_sub = ws1.max_row
        for col in range(1, max_cols + 1):
            cell = ws1.cell(row=fila_sub, column=col)
            cell.font = header_font
            cell.fill = sub_header_fill
            if col > 1: cell.alignment = Alignment(horizontal='right')

        def print_fin_row(label, key, is_bold=False):
            row_data = [label]
            for f in financials_sorted:
                row_data.append(f.get(key, 0.0))
            ws1.append(row_data)
            
            curr_row = ws1.max_row
            if is_bold:
                ws1.cell(row=curr_row, column=1).font = Font(bold=True)
                
            for col_idx in range(2, max_cols + 1):
                ws1.cell(row=curr_row, column=col_idx).style = currency_style

        # --- SECCIÓN BALANCE GENERAL ---
        ws1.append(["BALANCE GENERAL"] + [""] * len(years))
        ws1.cell(row=ws1.max_row, column=1).font = Font(bold=True, italic=True)
        
        print_fin_row("Activo", "input_assets")
        print_fin_row("Pasivo", "input_liabilities")
        print_fin_row("Capital Contable", "input_equity")
        
        ws1.append([]) # Separador

        # --- SECCIÓN ESTADO DE RESULTADOS ---
        ws1.append(["ESTADO DE RESULTADOS"] + [""] * len(years))
        ws1.cell(row=ws1.max_row, column=1).font = Font(bold=True, italic=True)

        print_fin_row("Ingresos Netos", "input_revenue")
        print_fin_row("Utilidad (Pérdida) Bruta", "input_gross_profit")
        print_fin_row("Utilidad (Pérdida) de Operación", "ebit")
        print_fin_row("Utilidad (Pérdida) antes de Impuestos", "ebt")
        print_fin_row("Impuestos a la Utilidad", "input_taxes")
        print_fin_row("Utilidad (Pérdida) Neta", "input_net_income", is_bold=True)

    # ==========================================
    # HOJA 2: FACTURACIÓN EN EL TIEMPO (RENOVADA)
    # ==========================================
    ws2 = wb.create_sheet("Facturación en el tiempo")
    
    # Matriz solicitada:
    # Cols: Last 24, 12, 9, 6, 3
    # Rows (Bloques): Revenue, Expenses, Inflows, Outflows, NFCF
    # Sub-Rows: Mean, Median Growth, Slope, CAGR
    
    periods_order = ["last_24_months", "last_12_months", "last_9_months", "last_6_months", "last_3_months"]
    metrics_map = ["revenue", "expenditures", "inflows", "outflows", "nfcf"]
    sub_metrics = [
        ("Promedio (Mean)", "mean", currency_style), # Media pura (moneda)
        ("Mediana (Median)", "median", currency_style), # Mediana pura (moneda)
        ("Crecimiento vs Periodo Anterior", "period_growth_rate", percent_style), # Crecimiento real (%)
        ("Slope (Tendencia Lineal)", "linear_slope", currency_style), # Pendiente de la tendencia (moneda)
        ("CAGR / CMGR", "cagr_cmgr", percent_style) # Crecimiento Compuesto (%)
    ]
    
    # Encabezados de Columnas
    ws2.append(["Métrica / Periodo"] + [p.replace("_", " ").title() for p in periods_order])
    estilo_header(ws2, 1)
    
    stats = data.get("stats_last_months", {})

    for metric_key in metrics_map:
        # Título de Sección (ej. REVENUE)
        ws2.append([metric_key.upper()])
        # Pintar fila de título sección
        curr_row = ws2.max_row
        ws2.cell(row=curr_row, column=1).fill = sub_header_fill
        ws2.cell(row=curr_row, column=1).font = header_font
        
        # Generar las 4 filas de sub-métricas
        for label, sub_key, style in sub_metrics:
            row_data = [label]
            for period in periods_order:
                # Navegar: stats -> period -> metric -> sub_metric
                # ej: stats['last_3_months']['revenue']['mean']
                p_data = stats.get(period) or {}
                m_data = p_data.get(metric_key) or {}
                val = m_data.get(sub_key, 0)
                row_data.append(val)
            
            ws2.append(row_data)
            
            # Aplicar estilos a las celdas de datos
            curr_row = ws2.max_row
            for col_idx in range(2, len(periods_order) + 2):
                cell = ws2.cell(row=curr_row, column=col_idx)
                cell.style = style

        # Espacio vacío
        ws2.append([""])

    ws2.column_dimensions['A'].width = 35
    for i in range(2, 7):
        ws2.column_dimensions[chr(64+i)].width = 20
    
    # ==========================================
    # HOJA 6: RED DE NEGOCIOS (CLIENTES Y PROVEEDORES)
    # ==========================================
    ws_net = wb.create_sheet("Red de Negocios")

    # --- FUNCIÓN PARA ANTIGÜEDAD DESDE RFC ---
    def calcular_antiguedad_rfc(rfc_str):
        if not rfc_str or not isinstance(rfc_str, str):
            return "N/A"
        
        rfc_str = rfc_str.strip().upper()
        current_year = datetime.now().year
        current_yy = current_year % 100 # ej. 2026 -> 26
        
        try:
            if len(rfc_str) == 12: # Persona Moral (3 letras + 6 números + 3 alfanuméricos)
                yy_str = rfc_str[3:5]
            elif len(rfc_str) == 13: # Persona Física (4 letras + 6 números + 3 alfanuméricos)
                yy_str = rfc_str[4:6]
            else:
                return "N/A"
                
            if not yy_str.isdigit():
                return "N/A"
                
            yy_int = int(yy_str)
            # Si el año del RFC es menor o igual al actual, es del 2000+. Sino, es 1900+
            year = (2000 + yy_int) if yy_int <= current_yy else (1900 + yy_int)
            return f"{current_year - year} Años"
        except:
            return "N/A"

    # --- FUNCIÓN PARA ARREGLAR PORCENTAJE ---
    def fix_percentage(val):
        try:
            v = float(val)
            # Si viene como 15.38, lo hacemos 0.1538 para que Excel lo formatee bien
            return v / 100 if v > 1 else v
        except:
            return 0.0

    # --- 1. CONCENTRACIÓN (TOP 5) ---
    concentration = data.get("concentration_last_12m", {})
    top_clients = concentration.get("top_5_clients", [])
    top_suppliers = concentration.get("top_5_suppliers", [])

    # === TOP 5 CLIENTES ===
    ws_net.append(["CONCENTRACIÓN: TOP 5 CLIENTES (Últimos 12 Meses)"])
    ws_net.merge_cells(start_row=ws_net.max_row, start_column=1, end_row=ws_net.max_row, end_column=5) # Ahora son 5 columnas
    estilo_header(ws_net, ws_net.max_row, 1, 5)
    
    ws_net.append(["Nombre", "RFC", "Monto Total", "Porcentaje", "Antigüedad / Edad"])
    sub_row = ws_net.max_row
    for col in range(1, 6):
        ws_net.cell(row=sub_row, column=col).font = header_font
        ws_net.cell(row=sub_row, column=col).fill = sub_header_fill
        ws_net.cell(row=sub_row, column=col).alignment = Alignment(horizontal='center')

    for c in top_clients:
        rfc = c.get("rfc", "")
        pct = fix_percentage(c.get("percentage", 0))
        antiguedad = calcular_antiguedad_rfc(rfc)
        
        ws_net.append([c.get("name"), rfc, c.get("total_amount"), pct, antiguedad])
        ws_net.cell(row=ws_net.max_row, column=3).style = currency_style
        ws_net.cell(row=ws_net.max_row, column=4).style = percent_style
        ws_net.cell(row=ws_net.max_row, column=5).alignment = Alignment(horizontal='center')

    ws_net.append([]) # Separador

    # === TOP 5 PROVEEDORES ===
    ws_net.append(["CONCENTRACIÓN: TOP 5 PROVEEDORES (Últimos 12 Meses)"])
    ws_net.merge_cells(start_row=ws_net.max_row, start_column=1, end_row=ws_net.max_row, end_column=5)
    estilo_header(ws_net, ws_net.max_row, 1, 5)
    
    ws_net.append(["Nombre", "RFC", "Monto Total", "Porcentaje", "Antigüedad / Edad"])
    sub_row = ws_net.max_row
    for col in range(1, 6):
        ws_net.cell(row=sub_row, column=col).font = header_font
        ws_net.cell(row=sub_row, column=col).fill = sub_header_fill
        ws_net.cell(row=sub_row, column=col).alignment = Alignment(horizontal='center')

    for s in top_suppliers:
        rfc = s.get("rfc", "")
        pct = fix_percentage(s.get("percentage", 0))
        antiguedad = calcular_antiguedad_rfc(rfc)
        
        ws_net.append([s.get("name"), rfc, s.get("total_amount"), pct, antiguedad])
        ws_net.cell(row=ws_net.max_row, column=3).style = currency_style
        ws_net.cell(row=ws_net.max_row, column=4).style = percent_style
        ws_net.cell(row=ws_net.max_row, column=5).alignment = Alignment(horizontal='center')

    ws_net.append([])
    ws_net.append([])

    # --- 2. DETALLE DE REDES (CUSTOMERS & VENDORS) ---
    networks = data.get("networks_data", {})
    customers_net = networks.get("customers", [])
    vendors_net = networks.get("vendors", [])

    # Columnas a imprimir para las redes
    net_headers = [
        "Nombre", "Total Recibido", "Total Cancelado", "% Cancelado", 
        "Descuentos", "Notas de Crédito", "Pago Pendiente", "Neto Recibido", 
        "Recibido PUE", "Recibido PPD", "Conteo PPD", "Monto Pagado", 
        "En Parcialidades", "Días Outstanding"
    ]

    def print_network_table(title, node_list):
        if not node_list: return
        
        ws_net.append([title])
        ws_net.merge_cells(start_row=ws_net.max_row, start_column=1, end_row=ws_net.max_row, end_column=len(net_headers))
        estilo_header(ws_net, ws_net.max_row, 1, len(net_headers))
        
        ws_net.append(net_headers)
        sub_row = ws_net.max_row
        for col in range(1, len(net_headers) + 1):
            ws_net.cell(row=sub_row, column=col).font = header_font
            ws_net.cell(row=sub_row, column=col).fill = sub_header_fill
            
        for n in node_list:
            ws_net.append([
                n.get("name"),
                n.get("total_received"),
                n.get("total_cancelled_received"),
                n.get("percentage_cancelled"),
                n.get("received_discounts"),
                n.get("received_credit_notes"),
                n.get("payment_pending"),
                n.get("net_received"),
                n.get("pue_received"),
                n.get("ppd_received"),
                n.get("ppd_count"),
                n.get("payment_amount"),
                n.get("in_installments"),
                n.get("days_outstanding")
            ])
            
            curr_row = ws_net.max_row
            # Aplicar formato Moneda a columnas correspondientes (B, C, E, F, G, H, I, J, L)
            # Índices en openpyxl: 2, 3, 5, 6, 7, 8, 9, 10, 12
            for col_idx in [2, 3, 5, 6, 7, 8, 9, 10, 12]:
                ws_net.cell(row=curr_row, column=col_idx).style = currency_style
                
            # Formato Porcentaje para % Cancelado (D -> 4)
            ws_net.cell(row=curr_row, column=4).style = percent_style
            
            # Formato numérico normal para Días y Parcialidades
            ws_net.cell(row=curr_row, column=13).number_format = '0.00'
            ws_net.cell(row=curr_row, column=14).number_format = '0.00'

        ws_net.append([])

    print_network_table("RED DETALLADA DE CLIENTES (CUSTOMER NETWORK)", customers_net)
    print_network_table("RED DETALLADA DE PROVEEDORES (VENDOR NETWORK)", vendors_net)

    # Ajuste de anchos para la Hoja de Redes
    ws_net.column_dimensions['A'].width = 45 # Nombres suelen ser muy largos
    for i in range(2, 15):
        ws_net.column_dimensions[chr(64+i)].width = 18

    # ==========================================
    # HOJA 4: PROYECCIONES (ACTUALIZADA)
    # ==========================================
    ws3 = wb.create_sheet("Proyecciones (Escenarios)")
    
    predictions = data.get("financial_predictions", {})
    metrics_to_export = ["revenue", "expenditures", "inflows", "outflows", "nfcf"]
    raw_history = data.get("raw_data_history", [])
    
    # 1. Obtener Fechas: 12 Pasadas + 12 Futuras
    # Extraemos fechas históricas (últimos 12 meses de la raw data)
    hist_dates_iso = [x["date"] for x in raw_history[-12:]] if raw_history else []
    
    # Extraemos fechas futuras del primer modelo
    future_dates_iso = []
    first_metric = predictions.get("revenue", {})
    if first_metric and first_metric.get("linear"):
        future_dates_iso = [pt["date"] for pt in first_metric["linear"]["scenarios"]["realistic"]]
    
    # Encabezados
    headers = ["Métrica", "Modelo", "Escenario", "Growth Comp.", "Growth Proy."] 
    # Añadir fechas históricas (Colapsadas visualmente o marcadas distinto)
    headers += [f"Hist: {d}" for d in hist_dates_iso]
    # Añadir fechas futuras
    headers += [f"Proy: {d}" for d in future_dates_iso]
    
    ws3.append(headers)
    estilo_header(ws3, 1)

    for metric in metrics_to_export:
        m_data = predictions.get(metric, {})
        if not m_data: continue
        
        # Obtener valores históricos para esta métrica (últimos 12)
        # Mapeamos nombre métrica a clave en raw_data_history
        metric_key_map = {
            "revenue": "revenue", "expenditures": "expenses",
            "inflows": "inflows_amount", "outflows": "outflows_amount",
            "nfcf": "nfcf"
        }
        key = metric_key_map.get(metric)
        hist_vals = [item.get(key, 0.0) for item in raw_history[-12:]]
        
        # Rellenar ceros si falta historia
        if len(hist_vals) < 12:
            hist_vals = [0.0]*(12-len(hist_vals)) + hist_vals

        for model_type in ["linear", "exponential", "seasonal"]:
            model_res = m_data.get(model_type)
            if not model_res: continue
            
            scenarios = model_res.get("scenarios", {})
            
            # Métricas de crecimiento
            # Comparison: Forecast Total vs History Total
            g_comp_r = model_res.get('growth_realistic') 
            g_comp_o = model_res.get('growth_optimistic')
            g_comp_p = model_res.get('growth_pessimistic')

            # Projection: Trend interna (campo nuevo agregado en Forecaster)
            g_proy_r = model_res.get('trend_realistic') 
            g_proy_o = model_res.get('trend_optimistic')
            g_proy_p = model_res.get('trend_pessimistic')
            
            # Función helper para formatear fila
            def make_row(label, comp, proy, futures):
                row = [
                    metric.upper() if label == "Realista" else "", # Métrica
                    model_type.title() if label == "Realista" else "", # Modelo
                    label, # Escenario
                    comp, # Growth Comparison
                    proy  # Growth Projection
                ]
                # Agregamos historia (Siempre la REAL para contexto visual)
                row += hist_vals 
                # Agregamos futuro
                row += [pt.get("value") for pt in futures]
                return row

            # Fila Realista
            ws3.append(make_row("Realista", g_comp_r, g_proy_r, scenarios.get("realistic", [])))
            # Fila Optimista
            ws3.append(make_row("Optimista (+)", g_comp_o, g_proy_o, scenarios.get("optimistic", [])))
            # Fila Pesimista
            ws3.append(make_row("Pesimista (-)", g_comp_p, g_proy_p, scenarios.get("pessimistic", [])))
            
            ws3.append([""]) # Separador visual

    # Estilos Proyecciones
    ws3.column_dimensions['A'].width = 15
    ws3.column_dimensions['B'].width = 15
    ws3.column_dimensions['C'].width = 20
    ws3.column_dimensions['D'].width = 15
    ws3.column_dimensions['E'].width = 15
    
    # Formato Porcentaje (Cols D y E)
    for row in ws3.iter_rows(min_row=2, min_col=4, max_col=5):
        for cell in row: cell.style = percent_style
        
    # Formato Moneda (Desde Col F hasta el final)
    max_col = ws3.max_column
    for row in ws3.iter_rows(min_row=2, min_col=6, max_col=max_col):
        for cell in row: 
            if isinstance(cell.value, (int, float)):
                cell.style = currency_style

    # Separador visual entre historia y futuro
    # La columna donde empieza el futuro es 6 (A-E) + 12 (hist) = 18?
    # A=1, B=2, C=3, D=4, E=5. Historia es 6 a 17. Futuro empieza en 18.
    col_split = 5 + len(hist_vals)
    for row in range(1, ws3.max_row + 1):
        cell = ws3.cell(row=row, column=col_split)
        cell.border = Border(right=Side(style='medium', color="000000"))


    # ==========================================
    # HOJA 5: RAW DATA (NUEVA PAGINA)
    # ==========================================
    ws_raw = wb.create_sheet("Raw Data")
    
    # Encabezados solicitados
    headers_raw = [
        "startDate", "Revenue", "Expenses", 
        "Inflows mxn amounts", "Outflows mxn amounts", 
        "NFCF", "Inflows transactions", "Outflows transactions"
    ]
    ws_raw.append(headers_raw)
    estilo_header(ws_raw, 1)
    
    for item in raw_history:
        # Item es un dict (pydantic model dump)
        ws_raw.append([
            item.get("date"),
            item.get("revenue"),
            item.get("expenses"),
            item.get("inflows_amount"),
            item.get("outflows_amount"),
            item.get("nfcf"),
            item.get("inflows_transactions"),
            item.get("outflows_transactions")
        ])

    # Formatos Raw Data
    ws_raw.column_dimensions['A'].width = 15
    for i in range(2, 9): ws_raw.column_dimensions[chr(64+i)].width = 20
    
    # Moneda para cols B a F
    for row in ws_raw.iter_rows(min_row=2, min_col=2, max_col=6):
        for cell in row: cell.style = currency_style

    # ==========================================
    # HOJA 6: ESTADOS FINANCIEROS Y ÁRBOL
    # ==========================================
    ws_fin = wb.create_sheet("Estados Financieros")
    
    financials = data.get("financial_ratios_history", [])
    
    if not financials:
        ws_fin.append(["No se encontró información de estados financieros."])
    else:
        # Aseguramos que los años estén ordenados de más reciente a más antiguo
        financials_sorted = sorted(financials, key=lambda x: str(x.get("year", "")), reverse=True)
        years = [str(f.get("year", "N/A")) for f in financials_sorted]
        
        # 1. Encabezados (Años)
        ws_fin.append(["Concepto / Año"] + years)
        estilo_header(ws_fin, 1, 1, len(years) + 1)
        
        # 2. Mapeo de Conceptos Financieros (Crudos y Razones)
        conceptos_financieros = [
            ("--- DATOS CONSOLIDADOS ---", None),
            ("Activos Totales", "input_assets"),
            ("Pasivo Total", "input_liabilities"),           
            ("Capital Contable (Equity)", "input_equity"),
            ("Ingresos Netos (Revenue)", "input_revenue"),
            ("Utilidad (Pérdida) Bruta", "input_gross_profit"),
            ("Utilidad Operativa (EBIT)", "ebit"),
            ("Utilidad antes de Impuestos (EBT)", "ebt"),
            ("Impuestos Calculados", "input_taxes"),
            ("Utilidad Neta Consolidada", "input_net_income"),
            ("NOPAT", "nopat"),
            
            ("", None), # Fila en blanco
            ("--- DATOS CRUDOS EXTRAÍDOS ---", None),
            ("Utilidad Neta (Crudo)", "raw_net_profit"),
            ("Pérdida Neta (Crudo)", "raw_net_loss"),
            ("Utilidad Operativa (Crudo)", "raw_ebit_profit"),
            ("Pérdida Operativa (Crudo)", "raw_ebit_loss"),
            ("Utilidad antes de Imp. (Crudo)", "raw_ebt_profit"),
            ("Pérdida antes de Imp. (Crudo)", "raw_ebt_loss"),
            
            ("", None), # Fila en blanco
            ("--- RAZONES FINANCIERAS ---", None),
            ("ROA", "roa"),
            ("ROE", "roe"),
            ("Margen Neto (%)", "net_profit_margin_percent")
        ]
        
        # 3. Construcción de Filas Iniciales
        for label, key in conceptos_financieros:
            row_data = [label]
            if key is None:
                for _ in financials_sorted: row_data.append("")
                ws_fin.append(row_data)
                ws_fin.cell(row=ws_fin.max_row, column=1).font = Font(bold=True)
                continue
                
            for f_year in financials_sorted:
                val = f_year.get(key, 0.0)
                row_data.append(val)
            ws_fin.append(row_data)
            
            curr_row = ws_fin.max_row
            for col_idx in range(2, len(years) + 2):
                cell = ws_fin.cell(row=curr_row, column=col_idx)
                if key in ["roa", "roe", "net_profit_margin_percent"]:
                    cell.style = percent_style
                    if key != "net_profit_margin_percent": 
                        cell.value = cell.value / 100 if cell.value else 0
                else:
                    cell.style = currency_style

        # Ajuste de Anchos de Columna iniciales
        ws_fin.column_dimensions['A'].width = 40
        for i in range(2, len(years) + 2):
            ws_fin.column_dimensions[chr(64+i)].width = 20

    # ---------------------------------------------------------
    # DESGLOSE COMPLETO DEL ÁRBOL FINANCIERO (Misma Hoja)
    # ---------------------------------------------------------
    fs_tree = data.get("financial_statements_tree", {})
    
    # 1. Función recursiva TOTALMENTE CIEGA para encontrar años
    all_years_tree = set()
    def extract_years_from_tree(node):
        if isinstance(node, dict):
            for k, v in node.items():
                if str(k).isdigit() and len(str(k)) == 4:
                    all_years_tree.add(str(k))
                elif isinstance(v, (dict, list)):
                    extract_years_from_tree(v)
        elif isinstance(node, list):
            for item in node:
                extract_years_from_tree(item)

    extract_years_from_tree(fs_tree)
    sorted_tree_years = sorted(list(all_years_tree), reverse=True)

    if sorted_tree_years:
        # 2. Función recursiva que escribe celda por celda y rastrea la fila actual
        def write_tree_node(ws, node, current_row, start_col, depth=0):
            if isinstance(node, dict) and "data" in node:
                return write_tree_node(ws, node["data"], current_row, start_col, depth)
            
            if isinstance(node, list):
                for item in node:
                    current_row = write_tree_node(ws, item, current_row, start_col, depth)
                return current_row
                
            if isinstance(node, dict):
                category = node.get("category", "")
                
                if category:
                    indent = "    " * depth
                    # Escribir categoría
                    ws.cell(row=current_row, column=start_col, value=f"{indent}{category}")
                    
                    if depth == 0:
                        ws.cell(row=current_row, column=start_col).font = Font(bold=True)
                        
                    # Escribir valores de los años
                    col_offset = 1
                    for y in sorted_tree_years:
                        val_node = node.get(y, 0.0)
                        val = 0.0
                        if isinstance(val_node, dict):
                            val = float(val_node.get("Total") or val_node.get("total") or 0.0)
                        elif isinstance(val_node, (int, float)):
                            val = float(val_node)
                        
                        c = ws.cell(row=current_row, column=start_col + col_offset, value=val)
                        c.style = currency_style
                        col_offset += 1
                        
                    current_row += 1
                
                children = node.get("children", [])
                if children:
                    next_depth = depth + 1 if category else depth
                    current_row = write_tree_node(ws, children, current_row, start_col, next_depth)
            
            return current_row

        # 3. Determinar coordenadas de inicio
        start_row_trees = ws_fin.max_row + 3
        col_bs = 1  # Balance Sheet empieza en columna A (1)
        col_is = len(sorted_tree_years) + 3  # Income Statement empieza dejando 1 columna vacía
        
        # --- ENCABEZADOS BALANCE ---
        ws_fin.cell(row=start_row_trees, column=col_bs, value="ESTADO DE POSICIÓN FINANCIERA")
        ws_fin.cell(row=start_row_trees, column=col_bs).font = Font(bold=True, size=12)
        
        ws_fin.cell(row=start_row_trees + 1, column=col_bs, value="Categoría")
        for i, y in enumerate(sorted_tree_years):
            ws_fin.cell(row=start_row_trees + 1, column=col_bs + i + 1, value=y)
        estilo_header(ws_fin, start_row_trees + 1, col_bs, col_bs + len(sorted_tree_years))
        
        # --- ENCABEZADOS RESULTADOS ---
        ws_fin.cell(row=start_row_trees, column=col_is, value="ESTADO DE RESULTADOS")
        ws_fin.cell(row=start_row_trees, column=col_is).font = Font(bold=True, size=12)
        
        ws_fin.cell(row=start_row_trees + 1, column=col_is, value="Categoría")
        for i, y in enumerate(sorted_tree_years):
            ws_fin.cell(row=start_row_trees + 1, column=col_is + i + 1, value=y)
        estilo_header(ws_fin, start_row_trees + 1, col_is, col_is + len(sorted_tree_years))
        
        # Ajustar anchos de columnas para el lado derecho
        col_letter_cat = ws_fin.cell(row=1, column=col_is).column_letter
        ws_fin.column_dimensions[col_letter_cat].width = 40
        for i in range(len(sorted_tree_years)):
            col_letter_year = ws_fin.cell(row=1, column=col_is + i + 1).column_letter
            ws_fin.column_dimensions[col_letter_year].width = 20

        # --- IMPRIMIR ÁRBOLES LADO A LADO ---
        data_start_row = start_row_trees + 2
        
        bs_tree = fs_tree.get("balance_sheet", {})
        write_tree_node(ws_fin, bs_tree, current_row=data_start_row, start_col=col_bs)
        
        is_tree = fs_tree.get("income_statement", {})
        write_tree_node(ws_fin, is_tree, current_row=data_start_row, start_col=col_is)

    # ==========================================
    # HOJA 7: DETALLE DE BURÓ (EXPANDIDO)
    # ==========================================
    ws4 = wb.create_sheet("Detalle de Buró")
    
    buro_info = data.get("buro_info", {})
    
    # ---------------------------------------------------------
    # Extraemos la data correcta sin importar qué nivel de anidación tenga
    # ---------------------------------------------------------
    raw_buro_container = buro_info.get("raw_buro_data", {})
    
    # A veces Syntage envuelve la data en otra llave "data" o "respuesta"
    raw_buro = raw_buro_container
    if "data" in raw_buro_container:
        raw_buro = raw_buro_container["data"]
    elif "respuesta" in raw_buro_container:
        raw_buro = raw_buro_container["respuesta"]
        
    # --- FUNCIONES AUXILIARES PARA TABLAS DINÁMICAS ---
    def build_kv_table(ws, title, dict_data):
        if not dict_data or not isinstance(dict_data, dict): return
        ws.append([title])
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=2)
        estilo_header(ws, ws.max_row, 1, 2)
        for k, v in dict_data.items():
            if not isinstance(v, (dict, list)): # Ignoramos anidados para mantenerlo plano
                ws.append([str(k), str(v)])
        ws.append([""]) # Separador

    def build_list_table(ws, title, list_data, cols_map):
        if not list_data: return
        
        # Fix: A veces (como en historialConsultas de empresas) viene como un solo dict, no como lista
        if isinstance(list_data, dict): 
            # Si el dict tiene las llaves adentro, lo envolvemos en lista
            if cols_map[0][1] in list_data:
                list_data = [list_data]
            else:
                # Si es un dict de dicts (menos común), extraemos los valores
                list_data = list(list_data.values())
                
        if not isinstance(list_data, list): return
        
        ws.append([title])
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=len(cols_map))
        estilo_header(ws, ws.max_row, 1, len(cols_map))
        
        headers = [c[0] for c in cols_map]
        ws.append(headers)
        sub_head_row = ws.max_row
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=sub_head_row, column=col)
            cell.font = header_font
            cell.fill = sub_header_fill
            cell.alignment = Alignment(horizontal='center')
            
        for item in list_data:
            if isinstance(item, dict):
                row = [str(item.get(c[1], "")) for c in cols_map]
                ws.append(row)
        ws.append([""]) # Separador

    # ---------------------------------------------------------
    # 1. DATOS GENERALES, CONSULTAS Y ENCABEZADO
    # ---------------------------------------------------------
    encabezado = raw_buro.get("encabezado", {})
    
    def format_buro_header_date(date_str):
        if not date_str: return ""
        d_str = str(date_str).strip()
        try:
            if "-" in d_str:
                d = datetime.strptime(d_str[:10], "%Y-%m-%d")
            elif len(d_str) == 8 and d_str.isdigit():
                d = datetime.strptime(d_str, "%d%m%Y")
            else:
                return d_str
            return f"{d.month}/{d.day}/{d.year}"
        except Exception:
            return d_str

    if "fechaConsulta" in encabezado:
        encabezado["fechaConsulta"] = format_buro_header_date(encabezado["fechaConsulta"])

    # 1. PINTAR ENCABEZADO
    build_kv_table(ws4, "ENCABEZADO / METADATOS", encabezado)
    
    # 2. PINTAR TABLA DE RESUMEN DE CONSULTAS (JUSTO EN MEDIO)
    inquiries_summary = buro_info.get("inquiries_summary", [])
    if inquiries_summary:
        ws4.append(["RESUMEN DE CONSULTAS (BÚSQUEDA DE CRÉDITO)"])
        ws4.merge_cells(start_row=ws4.max_row, start_column=1, end_row=ws4.max_row, end_column=5)
        estilo_header(ws4, ws4.max_row, 1, 5)
        
        headers_consultas = ["Concepto", "Cantidad", "Equivalente Meses", "Promedio Mensual", "Aumento vs Periodo Anterior"]
        ws4.append(headers_consultas)
        sub_head_row = ws4.max_row
        for col in range(1, 6):
            cell = ws4.cell(row=sub_head_row, column=col)
            cell.font = header_font
            cell.fill = sub_header_fill
            cell.alignment = Alignment(horizontal='center')
            
        for row in inquiries_summary:
            eq_months = row.get("equivalent_months")
            avg_month = row.get("monthly_average")
            growth = row.get("growth_vs_previous")
            
            ws4.append([
                row.get("concept"),
                row.get("quantity"),
                eq_months if eq_months is not None else "",
                avg_month if avg_month is not None else "",
                growth if growth is not None else ""
            ])
            
            curr_row = ws4.max_row
            if avg_month is not None:
                ws4.cell(row=curr_row, column=4).number_format = '0.0'
            if growth is not None:
                ws4.cell(row=curr_row, column=5).style = percent_style
                
        ws4.append([""]) # Fila separadora visual
        
    # 3. PINTAR DATOS GENERALES (EMPRESA) LIMPIOS (SIN DUPLICAR CONSULTAS)
    consultas_keys = [
        "consultaEmpresaComercialMas24Meses", "consultaEmpresaComercialUltimos24Meses",
        "consultaEmpresaComercialUltimos12Meses", "consultaEmpresaComercialUltimos3Meses",
        "consultaEntidadFinancieraMas24Meses", "consultaEntidadFinancieraUltimos24Meses",
        "consultaEntidadFinancieraUltimos12Meses", "consultaEntidadFinancieraUltimos3Meses"
    ]
    datos_generales_raw = raw_buro.get("datosGenerales", {})
    datos_generales_limpios = {k: v for k, v in datos_generales_raw.items() if k not in consultas_keys}
    build_kv_table(ws4, "DATOS GENERALES (EMPRESA)", datos_generales_limpios)


    # ---------------------------------------------------------
    # 2. SCORES DETALLADOS
    # ---------------------------------------------------------

    # Manejo si es Persona Física (PF)
    persona_node = raw_buro.get("persona", {})
    if not persona_node and "respuesta" in raw_buro:
        persona_node = raw_buro.get("respuesta", {}).get("persona", {})
        
    if persona_node:
        build_kv_table(ws4, "DATOS PERSONA (PF)", persona_node.get("nombre", {}))
        
        # Domicilios PF
        domicilios = persona_node.get("domicilios", {}).get("domicilio", [])
        build_list_table(ws4, "DOMICILIOS", domicilios, [
            ("Calle", "direccion1"), ("Colonia", "coloniaPoblacion"), ("Ciudad", "ciudad"), 
            ("Estado", "estado"), ("CP", "cp"), ("Fecha Registro", "fechaRegistroDomicilio")
        ])
        # Empleos PF
        empleos = persona_node.get("empleos", {}).get("empleo", [])
        build_list_table(ws4, "EMPLEOS", empleos, [
            ("Empresa", "nombreEmpresa"), ("Puesto", "puesto"), ("Salario", "salario"),
            ("Calle", "direccion1"), ("Colonia", "coloniaPoblacion"), ("Estado", "estado")
        ])
    
    score_list = raw_buro.get("score", [])
    if not score_list and "scoreBuroCredito" in persona_node:
        score_list = persona_node["scoreBuroCredito"]
    build_list_table(ws4, "SCORE DE CRÉDITO DETALLADO", score_list, [
        ("Nombre Score", "nombreScore"), ("Valor Score", "valorScore"), ("Código Score", "codigoScore"), 
        ("Razón 1", "codigoRazon1"), ("Razón 2", "codigoRazon2"), ("Razón 3", "codigoRazon3"), ("Error", "errorScore")
    ])

    # ---------------------------------------------------------
    # 3. ALERTAS (HAWK)
    # ---------------------------------------------------------
    hawk_list = raw_buro.get("hawkHr", [])
    build_list_table(ws4, "ALERTAS (HAWK)", hawk_list, [
        ("Código", "codigoHawk"), ("Fecha", "fechaMensajeHawk"), 
        ("Reporta", "tipoUsuarioReporta"), ("Descripción", "descripcionPrevencionHawk")
    ])

    # ---------------------------------------------------------
    # 4. PARÁMETROS DE CALIFICACIÓN
    # ---------------------------------------------------------
    califica_list = raw_buro.get("califica", [])
    build_list_table(ws4, "PARÁMETROS DE CALIFICACIÓN", califica_list, [
        ("Clave", "clave"), ("Nombre / Métrica", "nombre"), ("Valor", "valorCaracteristica")
    ])

    # ---------------------------------------------------------
    # 5. ACCIONISTAS (Empresas)
    # ---------------------------------------------------------
    accionistas_list = raw_buro.get("accionista", [])
    build_list_table(ws4, "ACCIONISTAS", accionistas_list, [
        ("Nombre", "nombreAccionista"), ("Paterno", "apellidoPaterno"), ("Materno", "apellidoMaterno"),
        ("RFC", "rfc"), ("CURP", "curp"), ("Porcentaje", "porcentaje"), ("Dirección", "direccion1")
    ])

    # ---------------------------------------------------------
    # 6. DESGLOSE COMPLETO DE CRÉDITO Y MOP
    # ---------------------------------------------------------
    ws4.append(["DESGLOSE DETALLADO DE CRÉDITOS"])
    ws4.merge_cells(start_row=ws4.max_row, start_column=1, end_row=ws4.max_row, end_column=32)
    estilo_header(ws4, ws4.max_row, 1, 32)
    
    lines = buro_info.get("credit_lines", [])
    
    # --- LÓGICA DE AGREGACIÓN PARA EL ENCABEZADO SUPERIOR ---
    # SUMIFS(S3:S25,M3:M25,"") -> Sumar saldo inicial si no hay fecha de cierre (evitando el string "None")
    sum_saldo_inicial_activo = sum(
        l.get("initial_balance", 0) for l in lines 
        if str(l.get("closing_date")).strip() in ["None", "N/A", ""]
    )
    # SUM(T)
    sum_saldo_vigente = sum(l.get("current_balance", 0) for l in lines)
    # SUM(V)
    sum_plazo_restante = sum(l.get("remaining_term_days", 0) for l in lines)
    # SUM(W)/360
    sum_pond_2 = sum(l.get("weighting_days", 0) for l in lines)
    pond_2_years = round(sum_pond_2 / 360, 2) if sum_pond_2 > 0 else 0.0
    # SUM(Y)
    sum_pago_mensual = sum(l.get("monthly_payment", 0) for l in lines)
    # =+T1/W1 (Pago anual)
    pago_anual = (sum_saldo_vigente / pond_2_years) if pond_2_years > 0 else 0.0
    # =+AA1/12 (Pago Mensual_2)
    pago_mensual_2 = pago_anual / 12

    # Construcción visual de la fila de agregados
    summary_1 = [""] * 31
    summary_1[9] = sum_saldo_inicial_activo
    summary_1[10] = sum_saldo_vigente
    summary_1[11] = 1.0 # 100% Ponderación
    summary_1[12] = sum_plazo_restante
    summary_1[13] = pond_2_years
    summary_1[14] = "años"
    summary_1[15] = sum_pago_mensual
    summary_1[16] = "Pago anual"
    summary_1[17] = pago_anual

    summary_2 = [""] * 31
    summary_2[16] = "Pago Mensual_2"
    summary_2[17] = pago_mensual_2

    ws4.append(summary_1)
    row_s1 = ws4.max_row
    ws4.append(summary_2)
    row_s2 = ws4.max_row

    # Estilos del bloque superior
    for col_idx in [10, 11, 16, 18]:
        ws4.cell(row=row_s1, column=col_idx).style = currency_style
    ws4.cell(row=row_s1, column=12).style = percent_style
    ws4.cell(row=row_s2, column=18).style = currency_style
    
    # Pintar celdas de título de Pago
    for r_idx in [row_s1, row_s2]:
        cell = ws4.cell(row=r_idx, column=17)
        cell.font = header_font
        cell.fill = header_fill

    # --- ENCABEZADOS DE LA TABLA PRINCIPAL ---
    headers_full = [
        "Número Cuenta", "Tipo Usuario", "Apertura", "Fecha Cierre", "Plazo", "Moneda", "Tipo Cambio", 
        "Atraso Mayor", "Tipo Crédito", "Saldo Inicial", "Saldo Vigente", "Ponderación", "Plazo Restante", 
        "Ponderación 2", "Fecha Final", "Pago Mensual", "Histórico Pagos", 
        "1", "2", "3", "4", "5", "6", "7", "9", "0", "U", "-", "LC", 
        "Último Actualizado", "Saldo Vencido"
    ]
    ws4.append(headers_full)
    sub_head_row = ws4.max_row
    for col in range(1, len(headers_full) + 1):
        cell = ws4.cell(row=sub_head_row, column=col)
        cell.font = header_font
        cell.fill = sub_header_fill
        cell.alignment = Alignment(horizontal='center')

    # Helper para fechas (convierte de YYYY-MM-DD a M/D/YYYY)
    def format_date_mdy(date_str):
        if not date_str or str(date_str).strip() in ["None", "N/A", ""]: return ""
        try:
            d = datetime.strptime(str(date_str)[:10], "%Y-%m-%d")
            return f"{d.month}/{d.day}/{d.year}"
        except:
            return str(date_str)
    
    # Helper rápido para convertir la fecha a formato "Jan-26"
    def format_period_mop(date_str):
        # Protegemos contra vacíos y el string literal "None"
        if not date_str or str(date_str).strip() == "None" or date_str == "N/A": 
            return "N/A"
        try:
            d = datetime.strptime(str(date_str)[:10], "%Y-%m-%d")
            return d.strftime("%b-%y").capitalize() # ej: Jan-26
        except Exception:
            return str(date_str)

    lines = buro_info.get("credit_lines", [])
    if not lines:
        ws4.append(["No se encontró información de créditos."])
    else:
        start_data_row = ws4.max_row + 1
        for line in lines:
            mop = line.get("mop_breakdown", {})
            hist = line.get("payment_history", "")
            
            ws4.append([
                line.get("account_number"),
                line.get("user_type"),
                format_date_mdy(line.get("opening_date")),
                format_date_mdy(line.get("closing_date")),
                line.get("term_days", 0),
                line.get("currency"),
                line.get("exchange_rate"),
                line.get("max_delay", 0),
                line.get("account_type"),
                line.get("initial_balance", 0.0),
                line.get("current_balance", 0.0),
                line.get("weighting_pct", 0.0),
                line.get("remaining_term_days", 0),
                line.get("weighting_days", 0.0),
                format_date_mdy(line.get("final_date")),
                line.get("monthly_payment", 0.0),
                hist,
                mop.get("mop_1", 0), mop.get("mop_2", 0), mop.get("mop_3", 0), mop.get("mop_4", 0),
                mop.get("mop_5", 0), mop.get("mop_6", 0), mop.get("mop_7", 0), mop.get("mop_9", 0),
                mop.get("mop_0", 0), mop.get("mop_u", 0), mop.get("mop_nd", 0), mop.get("mop_lc", 0),
                format_period_mop(line.get("update_date")),
                line.get("past_due_balance", 0.0)
            ])
            
            curr_row = ws4.max_row
            ws4.cell(row=curr_row, column=17).number_format = '@' # Histórico como texto
            
            # Estilos de celdas
            ws4.cell(row=curr_row, column=12).style = percent_style # Ponderación 1
            
            for col_idx in [10, 11, 16, 31]: # Saldos y Pagos a Moneda
                ws4.cell(row=curr_row, column=col_idx).style = currency_style 

    ws4.append([""])

    # ---------------------------------------------------------
    # 8. HISTORIAL DE CONSULTAS PROCESADAS
    # ---------------------------------------------------------
    ws4.append(["HISTORIAL DE CONSULTAS (CREDIT PULLS)"])
    ws4.merge_cells(start_row=ws4.max_row, start_column=1, end_row=ws4.max_row, end_column=4)
    estilo_header(ws4, ws4.max_row, 1, 4)
    
    headers_inq = ["Institución Solicitante", "Fecha Consulta", "Tipo Contrato", "Importe Solicitado"]
    ws4.append(headers_inq)
    sub_head_row = ws4.max_row
    for col in range(1, 5):
        cell = ws4.cell(row=sub_head_row, column=col)
        cell.font = header_font
        cell.fill = sub_header_fill
        cell.alignment = Alignment(horizontal='center')

    inquiries = buro_info.get("inquiries", [])
    if not inquiries:
        ws4.append(["No hay consultas recientes registradas."])
    else:
        start_data_row = ws4.max_row + 1
        for inq in inquiries:
            ws4.append([
                inq.get("institution"),
                inq.get("inquiry_date"),
                inq.get("contract_type"),
                inq.get("amount")
            ])
        # Formato moneda
        for row in ws4.iter_rows(min_row=start_data_row, min_col=4, max_col=4):
            for cell in row: cell.style = currency_style

    # Ajuste dinámico de anchos
    ws4.column_dimensions['A'].width = 35
    ws4.column_dimensions['B'].width = 20
    ws4.column_dimensions['C'].width = 25
    ws4.column_dimensions['D'].width = 20
    ws4.column_dimensions['E'].width = 15
    ws4.column_dimensions['F'].width = 15
    ws4.column_dimensions['G'].width = 15
    ws4.column_dimensions['H'].width = 15
    ws4.column_dimensions['I'].width = 28

    # ==========================================
    # GUARDAR
    # ==========================================
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()