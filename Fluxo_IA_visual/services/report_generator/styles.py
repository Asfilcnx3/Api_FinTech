# Fluxo_IA_visual/services/report_generator/styles.py
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# --- FUENTES Y RELLENOS PRINCIPALES ---
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
SUB_HEADER_FILL = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
ALERT_FILL = PatternFill(start_color="C0504D", end_color="C0504D", fill_type="solid")

# --- FORMATOS DE NÚMERO (Strings puros evitan bugs entre hojas en openpyxl) ---
CURRENCY_FORMAT = '$#,##0.00'
PERCENT_FORMAT = '0.00%'
DATE_FORMAT = 'YYYY-MM-DD'

# --- BORDES ---
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def aplicar_estilo_header(ws, row_idx, col_start=1, col_end=None):
    """Aplica el estilo de encabezado principal a un rango de celdas en una fila."""
    if col_end is None: 
        col_end = ws.max_column
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')

def aplicar_estilo_subheader(ws, row_idx, col_start=1, col_end=None):
    """Aplica el estilo de sub-encabezado a un rango de celdas en una fila."""
    if col_end is None:
        col_end = ws.max_column
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.font = HEADER_FONT
        cell.fill = SUB_HEADER_FILL
        cell.alignment = Alignment(horizontal='center')