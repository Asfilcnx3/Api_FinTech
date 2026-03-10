# Fluxo_IA_visual/services/report_generator/sheets/s05_financial_stmts.py
from openpyxl.styles import Font
from ..styles import (
    CURRENCY_FORMAT, PERCENT_FORMAT, 
    aplicar_estilo_header
)

def build(ws, data: dict):
    financials = data.get("financial_ratios_history", [])
    
    if not financials:
        ws.append(["No se encontró información de estados financieros."])
        return

    # 1. Ratios y Datos Consolidados
    financials_sorted = sorted(financials, key=lambda x: str(x.get("year", "")), reverse=True)
    years = [str(f.get("year", "N/A")) for f in financials_sorted]
    
    ws.append(["Concepto / Año"] + years)
    aplicar_estilo_header(ws, 1, 1, len(years) + 1)
    
    conceptos_financieros = [
        ("--- DATOS CONSOLIDADOS ---", None),
        ("Activos Totales", "input_assets"),
        ("Pasivo Total", "input_liabilities"),           
        ("Capital Contable (Equity)", "input_equity"),
        ("Ingresos Netos (Revenue)", "input_revenue"),
        ("Utilidad (Pérdida) Bruta", "input_gross_profit"), 
        ("Utilidad Operativa (EBIT)", "ebit"),
        ("Utilidad antes de Impuestos (EBT)", "ebt"),
        ("Impuestos Calculados", "input_taxes"),
        ("Utilidad Neta Consolidada", "input_net_income"),
        ("NOPAT", "nopat"),
        
        ("", None), 
        ("--- DATOS CRUDOS EXTRAÍDOS ---", None),
        ("Utilidad Neta (Crudo)", "raw_net_profit"),
        ("Pérdida Neta (Crudo)", "raw_net_loss"),
        ("Utilidad Operativa (Crudo)", "raw_ebit_profit"),
        ("Pérdida Operativa (Crudo)", "raw_ebit_loss"),
        ("Utilidad antes de Imp. (Crudo)", "raw_ebt_profit"),
        ("Pérdida antes de Imp. (Crudo)", "raw_ebt_loss"),
        
        ("", None), 
        ("--- RAZONES FINANCIERAS ---", None),
        ("ROA", "roa"),
        ("ROE", "roe"),
        ("Margen Neto (%)", "net_profit_margin_percent")
    ]
    
    for label, key in conceptos_financieros:
        row_data = [label]
        if key is None:
            for _ in financials_sorted: row_data.append("")
            ws.append(row_data)
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
            continue
            
        for f_year in financials_sorted:
            val = f_year.get(key, 0.0)
            row_data.append(val)
        ws.append(row_data)
        
        curr_row = ws.max_row
        for col_idx in range(2, len(years) + 2):
            cell = ws.cell(row=curr_row, column=col_idx)
            if key in ["roa", "roe", "net_profit_margin_percent"]:
                cell.number_format = PERCENT_FORMAT
                if key != "net_profit_margin_percent" and cell.value: 
                    cell.value = cell.value / 100 
            else:
                cell.number_format = CURRENCY_FORMAT

    ws.column_dimensions['A'].width = 40
    for i in range(2, len(years) + 2):
        ws.column_dimensions[chr(64+i)].width = 20

    # 2. ÁRBOL FINANCIERO (Recursivo)
    fs_tree = data.get("financial_statements_tree", {})
    all_years_tree = set()

    def extract_years_from_tree(node):
        if isinstance(node, dict):
            for k, v in node.items():
                if str(k).isdigit() and len(str(k)) == 4:
                    all_years_tree.add(str(k))
                elif isinstance(v, (dict, list)):
                    extract_years_from_tree(v)
        elif isinstance(node, list):
            for item in node:
                extract_years_from_tree(item)

    extract_years_from_tree(fs_tree)
    sorted_tree_years = sorted(list(all_years_tree), reverse=True)

    if sorted_tree_years:
        def write_tree_node(ws_obj, node, current_row, start_col, depth=0):
            if isinstance(node, dict) and "data" in node:
                return write_tree_node(ws_obj, node["data"], current_row, start_col, depth)
            
            if isinstance(node, list):
                for item in node:
                    current_row = write_tree_node(ws_obj, item, current_row, start_col, depth)
                return current_row
                
            if isinstance(node, dict):
                category = node.get("category", "")
                if category:
                    indent = "    " * depth
                    ws_obj.cell(row=current_row, column=start_col, value=f"{indent}{category}")
                    if depth == 0:
                        ws_obj.cell(row=current_row, column=start_col).font = Font(bold=True)
                        
                    col_offset = 1
                    for y in sorted_tree_years:
                        val_node = node.get(y, 0.0)
                        val = 0.0
                        if isinstance(val_node, dict):
                            val = float(val_node.get("Total") or val_node.get("total") or 0.0)
                        elif isinstance(val_node, (int, float)):
                            val = float(val_node)
                        
                        c = ws_obj.cell(row=current_row, column=start_col + col_offset, value=val)
                        c.number_format = CURRENCY_FORMAT
                        col_offset += 1
                        
                    current_row += 1
                
                children = node.get("children", [])
                if children:
                    next_depth = depth + 1 if category else depth
                    current_row = write_tree_node(ws_obj, children, current_row, start_col, next_depth)
            return current_row

        start_row_trees = ws.max_row + 3
        col_bs = 1  
        col_is = len(sorted_tree_years) + 3  
        
        # --- BALANCE ---
        ws.cell(row=start_row_trees, column=col_bs, value="ESTADO DE POSICIÓN FINANCIERA").font = Font(bold=True, size=12)
        ws.cell(row=start_row_trees + 1, column=col_bs, value="Categoría")
        for i, y in enumerate(sorted_tree_years):
            ws.cell(row=start_row_trees + 1, column=col_bs + i + 1, value=y)
        aplicar_estilo_header(ws, start_row_trees + 1, col_bs, col_bs + len(sorted_tree_years))
        
        # --- RESULTADOS ---
        ws.cell(row=start_row_trees, column=col_is, value="ESTADO DE RESULTADOS").font = Font(bold=True, size=12)
        ws.cell(row=start_row_trees + 1, column=col_is, value="Categoría")
        for i, y in enumerate(sorted_tree_years):
            ws.cell(row=start_row_trees + 1, column=col_is + i + 1, value=y)
        aplicar_estilo_header(ws, start_row_trees + 1, col_is, col_is + len(sorted_tree_years))
        
        col_letter_cat = ws.cell(row=1, column=col_is).column_letter
        ws.column_dimensions[col_letter_cat].width = 40
        for i in range(len(sorted_tree_years)):
            col_letter_year = ws.cell(row=1, column=col_is + i + 1).column_letter
            ws.column_dimensions[col_letter_year].width = 20

        data_start_row = start_row_trees + 2
        
        bs_tree = fs_tree.get("balance_sheet", {})
        write_tree_node(ws, bs_tree, current_row=data_start_row, start_col=col_bs)
        
        is_tree = fs_tree.get("income_statement", {})
        write_tree_node(ws, is_tree, current_row=data_start_row, start_col=col_is)