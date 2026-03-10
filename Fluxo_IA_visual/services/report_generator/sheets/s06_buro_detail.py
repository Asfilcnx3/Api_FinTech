# Fluxo_IA_visual/services/report_generator/sheets/s06_buro_detail.py
from openpyxl.styles import Alignment
from datetime import datetime
from ..styles import (
    HEADER_FONT, HEADER_FILL, SUB_HEADER_FILL, 
    CURRENCY_FORMAT, PERCENT_FORMAT, 
    aplicar_estilo_header
)

def build(ws, data: dict):
    buro_info = data.get("buro_info", {})
    raw_buro_container = buro_info.get("raw_buro_data", {})
    
    raw_buro = raw_buro_container
    if "data" in raw_buro_container:
        raw_buro = raw_buro_container["data"]
    elif "respuesta" in raw_buro_container:
        raw_buro = raw_buro_container["respuesta"]
        
    def build_kv_table(title, dict_data):
        if not dict_data or not isinstance(dict_data, dict): return
        ws.append([title])
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=2)
        aplicar_estilo_header(ws, ws.max_row, 1, 2)
        for k, v in dict_data.items():
            if not isinstance(v, (dict, list)): 
                ws.append([str(k), str(v)])
        ws.append([""])

    def build_list_table(title, list_data, cols_map):
        if not list_data: return
        if isinstance(list_data, dict): 
            if cols_map[0][1] in list_data:
                list_data = [list_data]
            else:
                list_data = list(list_data.values())
                
        if not isinstance(list_data, list): return
        
        ws.append([title])
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=len(cols_map))
        aplicar_estilo_header(ws, ws.max_row, 1, len(cols_map))
        
        headers = [c[0] for c in cols_map]
        ws.append(headers)
        sub_head_row = ws.max_row
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=sub_head_row, column=col)
            cell.font = HEADER_FONT
            cell.fill = SUB_HEADER_FILL
            cell.alignment = Alignment(horizontal='center')
            
        for item in list_data:
            if isinstance(item, dict):
                row = [str(item.get(c[1], "")) for c in cols_map]
                ws.append(row)
        ws.append([""])

    # 1. ENCABEZADOS Y CONSULTAS
    encabezado = raw_buro.get("encabezado", {})
    build_kv_table("ENCABEZADO / METADATOS", encabezado)
    
    inquiries_summary = buro_info.get("inquiries_summary", [])
    if inquiries_summary:
        ws.append(["RESUMEN DE CONSULTAS (BÚSQUEDA DE CRÉDITO)"])
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=5)
        aplicar_estilo_header(ws, ws.max_row, 1, 5)
        
        headers_consultas = ["Concepto", "Cantidad", "Equivalente Meses", "Promedio Mensual", "Aumento vs Periodo Anterior"]
        ws.append(headers_consultas)
        sub_head_row = ws.max_row
        for col in range(1, 6):
            cell = ws.cell(row=sub_head_row, column=col)
            cell.font = HEADER_FONT
            cell.fill = SUB_HEADER_FILL
            cell.alignment = Alignment(horizontal='center')
            
        for row in inquiries_summary:
            avg_month = row.get("monthly_average")
            growth = row.get("growth_vs_previous")
            ws.append([
                row.get("concept"), row.get("quantity"),
                row.get("equivalent_months") if row.get("equivalent_months") is not None else "",
                avg_month if avg_month is not None else "",
                growth if growth is not None else ""
            ])
            curr_row = ws.max_row
            if avg_month is not None: ws.cell(row=curr_row, column=4).number_format = '0.0'
            if growth is not None: ws.cell(row=curr_row, column=5).number_format = PERCENT_FORMAT
        ws.append([""]) 
        
    consultas_keys = [
        "consultaEmpresaComercialMas24Meses", "consultaEmpresaComercialUltimos24Meses",
        "consultaEmpresaComercialUltimos12Meses", "consultaEmpresaComercialUltimos3Meses",
        "consultaEntidadFinancieraMas24Meses", "consultaEntidadFinancieraUltimos24Meses",
        "consultaEntidadFinancieraUltimos12Meses", "consultaEntidadFinancieraUltimos3Meses"
    ]
    datos_generales_raw = raw_buro.get("datosGenerales", {})
    datos_generales_limpios = {k: v for k, v in datos_generales_raw.items() if k not in consultas_keys}
    build_kv_table("DATOS GENERALES (EMPRESA)", datos_generales_limpios)

    # 2. SECCIONES SECUNDARIAS
    persona_node = raw_buro.get("persona", {})
    if not persona_node and "respuesta" in raw_buro:
        persona_node = raw_buro.get("respuesta", {}).get("persona", {})
        
    if persona_node:
        build_kv_table("DATOS PERSONA (PF)", persona_node.get("nombre", {}))
        domicilios = persona_node.get("domicilios", {}).get("domicilio", [])
        build_list_table("DOMICILIOS", domicilios, [
            ("Calle", "direccion1"), ("Colonia", "coloniaPoblacion"), ("Ciudad", "ciudad"), 
            ("Estado", "estado"), ("CP", "cp"), ("Fecha Registro", "fechaRegistroDomicilio")
        ])
        empleos = persona_node.get("empleos", {}).get("empleo", [])
        build_list_table("EMPLEOS", empleos, [
            ("Empresa", "nombreEmpresa"), ("Puesto", "puesto"), ("Salario", "salario"),
            ("Calle", "direccion1"), ("Colonia", "coloniaPoblacion"), ("Estado", "estado")
        ])
    
    score_list = raw_buro.get("score", [])
    if not score_list and "scoreBuroCredito" in persona_node:
        score_list = persona_node["scoreBuroCredito"]
    build_list_table("SCORE DE CRÉDITO DETALLADO", score_list, [
        ("Nombre Score", "nombreScore"), ("Valor Score", "valorScore"), ("Código Score", "codigoScore"), 
        ("Razón 1", "codigoRazon1"), ("Razón 2", "codigoRazon2"), ("Razón 3", "codigoRazon3"), ("Error", "errorScore")
    ])

    build_list_table("ALERTAS (HAWK)", raw_buro.get("hawkHr", []), [
        ("Código", "codigoHawk"), ("Fecha", "fechaMensajeHawk"), 
        ("Reporta", "tipoUsuarioReporta"), ("Descripción", "descripcionPrevencionHawk")
    ])

    build_list_table("PARÁMETROS DE CALIFICACIÓN", raw_buro.get("califica", []), [
        ("Clave", "clave"), ("Nombre / Métrica", "nombre"), ("Valor", "valorCaracteristica")
    ])

    build_list_table("ACCIONISTAS", raw_buro.get("accionista", []), [
        ("Nombre", "nombreAccionista"), ("Paterno", "apellidoPaterno"), ("Materno", "apellidoMaterno"),
        ("RFC", "rfc"), ("CURP", "curp"), ("Porcentaje", "porcentaje"), ("Dirección", "direccion1")
    ])

    # 3. LÍNEAS DE CRÉDITO
    ws.append(["DESGLOSE DETALLADO DE CRÉDITOS"])
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=32)
    aplicar_estilo_header(ws, ws.max_row, 1, 32)
    
    lines = buro_info.get("credit_lines", [])
    
    sum_saldo_inicial_activo = sum(l.get("initial_balance", 0) for l in lines if str(l.get("closing_date")).strip() in ["None", "N/A", ""])
    sum_saldo_vigente = sum(l.get("current_balance", 0) for l in lines)
    sum_plazo_restante = sum(l.get("remaining_term_days", 0) for l in lines)
    sum_pond_2 = sum(l.get("weighting_days", 0) for l in lines)
    pond_2_years = round(sum_pond_2 / 360, 2) if sum_pond_2 > 0 else 0.0
    sum_pago_mensual = sum(l.get("monthly_payment", 0) for l in lines)
    pago_anual = (sum_saldo_vigente / pond_2_years) if pond_2_years > 0 else 0.0
    pago_mensual_2 = pago_anual / 12

    summary_1 = [""] * 31
    summary_1[9] = sum_saldo_inicial_activo
    summary_1[10] = sum_saldo_vigente
    summary_1[11] = 1.0 
    summary_1[12] = sum_plazo_restante
    summary_1[13] = pond_2_years
    summary_1[14] = "años"
    summary_1[15] = sum_pago_mensual
    summary_1[16] = "Pago anual"
    summary_1[17] = pago_anual

    summary_2 = [""] * 31
    summary_2[16] = "Pago Mensual_2"
    summary_2[17] = pago_mensual_2

    ws.append(summary_1)
    row_s1 = ws.max_row
    ws.append(summary_2)
    row_s2 = ws.max_row

    for col_idx in [10, 11, 16, 18]:
        ws.cell(row=row_s1, column=col_idx).number_format = CURRENCY_FORMAT
    ws.cell(row=row_s1, column=12).number_format = PERCENT_FORMAT
    ws.cell(row=row_s2, column=18).number_format = CURRENCY_FORMAT
    
    for r_idx in [row_s1, row_s2]:
        cell = ws.cell(row=r_idx, column=17)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    headers_full = [
        "Número Cuenta", "Tipo Usuario", "Apertura", "Fecha Cierre", "Plazo", "Moneda", "Tipo Cambio", 
        "Atraso Mayor", "Tipo Crédito", "Saldo Inicial", "Saldo Vigente", "Ponderación", "Plazo Restante", 
        "Ponderación 2", "Fecha Final", "Pago Mensual", "Histórico Pagos", 
        "1", "2", "3", "4", "5", "6", "7", "9", "0", "U", "-", "LC", 
        "Último Actualizado", "Saldo Vencido"
    ]
    ws.append(headers_full)
    sub_head_row = ws.max_row
    for col in range(1, len(headers_full) + 1):
        cell = ws.cell(row=sub_head_row, column=col)
        cell.font = HEADER_FONT
        cell.fill = SUB_HEADER_FILL
        cell.alignment = Alignment(horizontal='center')

    def format_date_mdy(date_str):
        if not date_str or str(date_str).strip() in ["None", "N/A", ""]: return ""
        try:
            d = datetime.strptime(str(date_str)[:10], "%Y-%m-%d")
            return f"{d.month}/{d.day}/{d.year}"
        except: return str(date_str)
    
    def format_period_mop(date_str):
        if not date_str or str(date_str).strip() in ["None", "N/A", ""]: return "N/A"
        try:
            d = datetime.strptime(str(date_str)[:10], "%Y-%m-%d")
            return d.strftime("%b-%y").capitalize()
        except: return str(date_str)

    if not lines:
        ws.append(["No se encontró información de créditos."])
    else:
        for line in lines:
            mop = line.get("mop_breakdown", {})
            hist = line.get("payment_history", "")
            
            ws.append([
                line.get("account_number"), line.get("user_type"), format_date_mdy(line.get("opening_date")),
                format_date_mdy(line.get("closing_date")), line.get("term_days", 0), line.get("currency"),
                line.get("exchange_rate"), line.get("max_delay", 0), line.get("account_type"),
                line.get("initial_balance", 0.0), line.get("current_balance", 0.0), line.get("weighting_pct", 0.0),
                line.get("remaining_term_days", 0), line.get("weighting_days", 0.0), format_date_mdy(line.get("final_date")),
                line.get("monthly_payment", 0.0), hist,
                mop.get("mop_1", 0), mop.get("mop_2", 0), mop.get("mop_3", 0), mop.get("mop_4", 0),
                mop.get("mop_5", 0), mop.get("mop_6", 0), mop.get("mop_7", 0), mop.get("mop_9", 0),
                mop.get("mop_0", 0), mop.get("mop_u", 0), mop.get("mop_nd", 0), mop.get("mop_lc", 0),
                format_period_mop(line.get("update_date")), line.get("past_due_balance", 0.0)
            ])
            
            curr_row = ws.max_row
            ws.cell(row=curr_row, column=17).number_format = '@' 
            ws.cell(row=curr_row, column=12).number_format = PERCENT_FORMAT 
            for col_idx in [10, 11, 16, 31]: 
                ws.cell(row=curr_row, column=col_idx).number_format = CURRENCY_FORMAT 

    ws.append([""])

    # 4. HISTORIAL DE CONSULTAS
    ws.append(["HISTORIAL DE CONSULTAS (CREDIT PULLS)"])
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=4)
    aplicar_estilo_header(ws, ws.max_row, 1, 4)
    
    headers_inq = ["Institución Solicitante", "Fecha Consulta", "Tipo Contrato", "Importe Solicitado"]
    ws.append(headers_inq)
    sub_head_row = ws.max_row
    for col in range(1, 5):
        cell = ws.cell(row=sub_head_row, column=col)
        cell.font = HEADER_FONT
        cell.fill = SUB_HEADER_FILL
        cell.alignment = Alignment(horizontal='center')

    inquiries = buro_info.get("inquiries", [])
    if not inquiries:
        ws.append(["No hay consultas recientes registradas."])
    else:
        start_data_row = ws.max_row + 1
        for inq in inquiries:
            ws.append([
                inq.get("institution"), inq.get("inquiry_date"),
                inq.get("contract_type"), inq.get("amount")
            ])
        for row in ws.iter_rows(min_row=start_data_row, min_col=4, max_col=4):
            for cell in row: cell.number_format = CURRENCY_FORMAT

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 28