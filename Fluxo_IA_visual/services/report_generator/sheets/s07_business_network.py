# Fluxo_IA_visual/services/report_generator/sheets/s07_business_network.py
from openpyxl.styles import Alignment
from datetime import datetime
from ..styles import (
    HEADER_FONT, SUB_HEADER_FILL, 
    CURRENCY_FORMAT, PERCENT_FORMAT, 
    aplicar_estilo_header, Font
)

def calcular_antiguedad_rfc(rfc_str):
    if not rfc_str or not isinstance(rfc_str, str): return "N/A"
    rfc_str = rfc_str.strip().upper()
    current_year = datetime.now().year
    current_yy = current_year % 100 
    try:
        if len(rfc_str) == 12: yy_str = rfc_str[3:5]
        elif len(rfc_str) == 13: yy_str = rfc_str[4:6]
        else: return "N/A"
        if not yy_str.isdigit(): return "N/A"
        yy_int = int(yy_str)
        year = (2000 + yy_int) if yy_int <= current_yy else (1900 + yy_int)
        return f"{current_year - year}"
    except: return "N/A"

def fix_percentage(val):
    try:
        v = float(val)
        return v / 100 if v > 1 else v
    except: return 0.0

def build(ws, data: dict):
    concentration = data.get("concentration_last_12m", {})
    top_clients = concentration.get("top_5_clients", [])
    top_suppliers = concentration.get("top_5_suppliers", [])

    def pintar_concentracion(titulo, datos):
        ws.append([titulo])
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=7) # <-- Expandido a 7 columnas
        aplicar_estilo_header(ws, ws.max_row, 1, 7)
        
        ws.append(["Nombre", "RFC", "Monto Total", "Porcentaje", "Antigüedad / Edad", "Slope (Tendencia)", "Comportamiento"])
        sub_row = ws.max_row
        for col in range(1, 8):
            cell = ws.cell(row=sub_row, column=col)
            cell.font = HEADER_FONT
            cell.fill = SUB_HEADER_FILL
            cell.alignment = Alignment(horizontal='center')

        for item in datos:
            rfc = item.get("rfc", "")
            pct = fix_percentage(item.get("percentage", 0))
            antiguedad = calcular_antiguedad_rfc(rfc)
            slope = item.get("linear_slope", 0.0)
            trend = item.get("trend_text", "N/A")
            
            ws.append([item.get("name"), rfc, item.get("total_amount"), pct, antiguedad, slope, trend])
            
            curr_row = ws.max_row
            ws.cell(row=curr_row, column=3).number_format = CURRENCY_FORMAT
            ws.cell(row=curr_row, column=4).number_format = PERCENT_FORMAT
            ws.cell(row=curr_row, column=5).alignment = Alignment(horizontal='center')
            
            # Formato moneda para la pendiente (representa crecimiento en $ por mes)
            ws.cell(row=curr_row, column=6).number_format = CURRENCY_FORMAT
            ws.cell(row=curr_row, column=7).alignment = Alignment(horizontal='center')
            
            # Semáforo visual para el comportamiento
            if trend == "Creciendo":
                ws.cell(row=curr_row, column=7).font = Font(color="006100", bold=True) # Verde
            elif trend == "Disminuyendo":
                ws.cell(row=curr_row, column=7).font = Font(color="9C0006", bold=True) # Rojo
                
        ws.append([])

    pintar_concentracion("CONCENTRACIÓN: TOP 5 CLIENTES (Últimos 12 Meses)", top_clients)
    pintar_concentracion("CONCENTRACIÓN: TOP 5 PROVEEDORES (Últimos 12 Meses)", top_suppliers)
    ws.append([])

    networks = data.get("networks_data", {})
    customers_net = networks.get("customers", [])
    vendors_net = networks.get("vendors", [])

    net_headers = [
        "Nombre", "Total Recibido", "Total Cancelado", "% Cancelado", 
        "Descuentos", "Notas de Crédito", "Pago Pendiente", "Neto Recibido", 
        "Recibido PUE", "Recibido PPD", "Conteo PPD", "Monto Pagado", 
        "En Parcialidades", "Días Outstanding"
    ]

    def print_network_table(title, node_list):
        if not node_list: return
        ws.append([title])
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=len(net_headers))
        aplicar_estilo_header(ws, ws.max_row, 1, len(net_headers))
        
        ws.append(net_headers)
        sub_row = ws.max_row
        for col in range(1, len(net_headers) + 1):
            cell = ws.cell(row=sub_row, column=col)
            cell.font = HEADER_FONT
            cell.fill = SUB_HEADER_FILL
            
        for n in node_list:
            ws.append([
                n.get("name"), n.get("total_received"), n.get("total_cancelled_received"),
                fix_percentage(n.get("percentage_cancelled", 0)), # También usamos el fix_percentage aquí por seguridad
                n.get("received_discounts"), n.get("received_credit_notes"), n.get("payment_pending"),
                n.get("net_received"), n.get("pue_received"), n.get("ppd_received"),
                n.get("ppd_count"), n.get("payment_amount"), n.get("in_installments"), n.get("days_outstanding")
            ])
            
            curr_row = ws.max_row
            for col_idx in [2, 3, 5, 6, 7, 8, 9, 10, 12]:
                ws.cell(row=curr_row, column=col_idx).number_format = CURRENCY_FORMAT
                
            ws.cell(row=curr_row, column=4).number_format = PERCENT_FORMAT
            ws.cell(row=curr_row, column=13).number_format = '0.00'
            ws.cell(row=curr_row, column=14).number_format = '0.00'
        ws.append([])

    print_network_table("RED DETALLADA DE CLIENTES (CUSTOMER NETWORK)", customers_net)
    print_network_table("RED DETALLADA DE PROVEEDORES (VENDOR NETWORK)", vendors_net)

    ws.column_dimensions['A'].width = 45 
    for i in range(2, 15):
        ws.column_dimensions[chr(64+i)].width = 18