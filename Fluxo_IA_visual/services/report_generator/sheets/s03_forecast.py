# Fluxo_IA_visual/services/report_generator/sheets/s03_forecast.py
from openpyxl.styles import Border, Side
from ..styles import (
    CURRENCY_FORMAT, PERCENT_FORMAT, 
    aplicar_estilo_header
)

def build(ws, data: dict):
    predictions = data.get("financial_predictions", {})
    metrics_to_export = ["revenue", "expenditures", "inflows", "outflows", "nfcf"]
    raw_history = data.get("raw_data_history", [])
    
    # 1. Obtener Fechas: 12 Pasadas + 12 Futuras
    hist_dates_iso = [x.get("date") for x in raw_history[-12:]] if raw_history else []
    
    future_dates_iso = []
    first_metric = predictions.get("revenue", {})
    if first_metric and first_metric.get("linear"):
        future_dates_iso = [pt.get("date") for pt in first_metric["linear"]["scenarios"]["realistic"]]
    
    # Encabezados
    headers = ["Métrica", "Modelo", "Escenario", "Growth Comp.", "Growth Proy."] 
    headers += [f"Hist: {d}" for d in hist_dates_iso]
    headers += [f"Proy: {d}" for d in future_dates_iso]
    
    ws.append(headers)
    aplicar_estilo_header(ws, 1)

    for metric in metrics_to_export:
        m_data = predictions.get(metric, {})
        if not m_data: continue
        
        metric_key_map = {
            "revenue": "revenue", "expenditures": "expenses",
            "inflows": "inflows_amount", "outflows": "outflows_amount",
            "nfcf": "nfcf"
        }
        key = metric_key_map.get(metric)
        hist_vals = [item.get(key, 0.0) for item in raw_history[-12:]]
        
        # Rellenar ceros si falta historia
        if len(hist_vals) < 12:
            hist_vals = [0.0]*(12-len(hist_vals)) + hist_vals

        for model_type in ["linear", "exponential", "seasonal"]:
            model_res = m_data.get(model_type)
            if not model_res: continue
            
            scenarios = model_res.get("scenarios", {})
            g_comp_r = model_res.get('comparison_realistic') 
            g_comp_o = model_res.get('comparison_optimistic')
            g_comp_p = model_res.get('comparison_pessimistic')

            g_proy_r = model_res.get('trend_realistic') 
            g_proy_o = model_res.get('trend_optimistic')
            g_proy_p = model_res.get('trend_pessimistic')
            
            def make_row(label, comp, proy, futures):
                row = [
                    metric.upper() if label == "Realista" else "",
                    model_type.title() if label == "Realista" else "",
                    label, comp, proy
                ]
                row += hist_vals 
                row += [pt.get("value") for pt in futures]
                return row

            ws.append(make_row("Realista", g_comp_r, g_proy_r, scenarios.get("realistic", [])))
            ws.append(make_row("Optimista (+)", g_comp_o, g_proy_o, scenarios.get("optimistic", [])))
            ws.append(make_row("Pesimista (-)", g_comp_p, g_proy_p, scenarios.get("pessimistic", [])))
            
            ws.append([""]) # Separador visual

    # --- ESTILOS DE LA HOJA ---
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    # Formato Porcentaje (Cols D y E)
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=5):
        for cell in row: 
            cell.number_format = PERCENT_FORMAT
            
    # Formato Moneda (Desde Col F hasta el final)
    max_col = ws.max_column
    for row in ws.iter_rows(min_row=2, min_col=6, max_col=max_col):
        for cell in row: 
            if isinstance(cell.value, (int, float)):
                cell.number_format = CURRENCY_FORMAT

    # Borde separador visual entre historia y futuro
    col_split = 5 + len(hist_vals)
    for row in range(1, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_split)
        # Mantiene el borde derecho para marcar dónde empieza el futuro
        cell.border = Border(right=Side(style='medium', color="000000"))