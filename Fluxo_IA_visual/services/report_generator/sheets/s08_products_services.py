# Fluxo_IA_visual/services/report_generator/sheets/s08_products_services.py
from openpyxl.styles import Alignment, Font
import re
from ..styles import (
    HEADER_FONT, SUB_HEADER_FILL, 
    CURRENCY_FORMAT, PERCENT_FORMAT, 
    aplicar_estilo_header
)

def fix_percentage(val):
    try:
        v = float(val)
        return v / 100 if v > 1 else v
    except: return 0.0

def build(ws, data: dict):
    # 1. ACTIVIDADES ECONÓMICAS
    ws.append(["ACTIVIDADES ECONÓMICAS (GIRO REGISTRADO SAT)"])
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=3)
    aplicar_estilo_header(ws, ws.max_row, 1, 3)
    
    ws.append(["Actividad", "Porcentaje Ingresos", "Fecha Inicio"])
    sub_row = ws.max_row
    for col in range(1, 4):
        ws.cell(row=sub_row, column=col).font = HEADER_FONT
        ws.cell(row=sub_row, column=col).fill = SUB_HEADER_FILL
        
    acts = data.get("economic_activities", [])
    ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

    for act in acts:
        # 1. Obtenemos y limpiamos el texto
        raw_name = str(act.get("name", "N/A"))
        clean_name = ILLEGAL_CHARACTERS_RE.sub("", raw_name)
        clean_name = " ".join(clean_name.split())
        
        # 2. Insertamos a la celda usando el texto limpio
        ws.append([clean_name, fix_percentage(act.get("percentage", 0)), act.get("start_date")])
        ws.cell(row=ws.max_row, column=2).number_format = PERCENT_FORMAT

    ws.append([])
    ws.append([])

    # --- 2. ANÁLISIS IA ---
    products_data = data.get("products_data", {})
    if products_data:
        ws.append(["ANÁLISIS DE CONGRUENCIA Y TENDENCIAS (INTELIGENCIA ARTIFICIAL)"])
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=5)
        aplicar_estilo_header(ws, ws.max_row, 1, 5)

        # A. Red Flags y Conceptos Genéricos
        ws.append(["1. Congruencia General y Alertas (Red Flags / Conceptos Genéricos):"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, color="000000")
        
        ws.append([products_data.get("llm_activity_analysis", "Sin análisis disponible.")])
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=5)
        ws.cell(row=ws.max_row, column=1).alignment = Alignment(wrap_text=True, vertical='top')
        ws.row_dimensions[ws.max_row].height = 60  

        ws.append([])

        # B. Análisis de Ventas (Peso vs CSF)
        ws.append(["2. Análisis de Ventas (Peso real vs CSF y Tendencia Operativa):"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, color="000000")

        ws.append([products_data.get("llm_ventas_peso_tendencia", "Sin análisis disponible.")])
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=5)
        ws.cell(row=ws.max_row, column=1).alignment = Alignment(wrap_text=True, vertical='top')
        ws.row_dimensions[ws.max_row].height = 60  

        ws.append([])

        # C. Análisis de Compras (Insumos Clave)
        ws.append(["3. Análisis de Compras (Insumos Clave y Tendencia de Gasto):"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, color="000000")

        ws.append([products_data.get("llm_compras_insumos", "Sin análisis disponible.")])
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=5)
        ws.cell(row=ws.max_row, column=1).alignment = Alignment(wrap_text=True, vertical='top')
        ws.row_dimensions[ws.max_row].height = 60  

    ws.append([])
    ws.append([])

    # 3. PRODUCTOS Y SERVICIOS
    sold = products_data.get("sold", [])
    bought = products_data.get("bought", [])

    headers_ps = ["Descripción / Concepto", "Monto Total", "Porcentaje (Share)", "Transacciones"]

    def pintar_tabla_productos(titulo, lista_items):
        if not lista_items: return
        ws.append([titulo])
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=4)
        aplicar_estilo_header(ws, ws.max_row, 1, 4)
        
        ws.append(headers_ps)
        curr_sub = ws.max_row
        for col in range(1, 5):
            cell = ws.cell(row=curr_sub, column=col)
            cell.font = HEADER_FONT
            cell.fill = SUB_HEADER_FILL
            cell.alignment = Alignment(horizontal='center')
            
        for item in lista_items:
            ws.append([
                item.get("description"),
                item.get("total_amount"),
                fix_percentage(item.get("percentage", 0)),
                item.get("transactions")
            ])
            r = ws.max_row
            ws.cell(row=r, column=2).number_format = CURRENCY_FORMAT
            ws.cell(row=r, column=3).number_format = PERCENT_FORMAT
            ws.cell(row=r, column=4).alignment = Alignment(horizontal='center')
        ws.append([])

    pintar_tabla_productos("TOP 50: PRODUCTOS Y SERVICIOS VENDIDOS (Últimos 12 Meses)", sold)
    pintar_tabla_productos("TOP 50: PRODUCTOS Y SERVICIOS COMPRADOS (Últimos 12 Meses)", bought)

    # Anchos de columna
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15