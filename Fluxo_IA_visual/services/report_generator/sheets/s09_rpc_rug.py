from openpyxl.styles import Alignment, Font
from ..styles import (
    aplicar_estilo_header, aplicar_estilo_subheader, CURRENCY_FORMAT
)

def build(ws, data: dict):
    # Recuerda que `data` viene de un dict, así que usamos .get()
    registry_data = data.get("registry_data") or {}
    
    rpc_records = registry_data.get("rpc_records", [])
    rug_records = registry_data.get("rug_records", [])
    
    # --- SECCIÓN RPC ---
    ws.append(["REGISTRO PÚBLICO DE COMERCIO (RPC)"])
    ws.merge_cells('A1:D1')
    aplicar_estilo_header(ws, 1, 1, 4)
    
    if rpc_records:
        for r in rpc_records:
            # 1. Datos Base
            ws.append(["DATOS DE LA SOCIEDAD"])
            ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=4)
            aplicar_estilo_subheader(ws, ws.max_row, 1, 4)
            
            ws.append(["Folio Mercantil", "Fecha de Registro", "Entidad Federativa", "Razón Social"])
            for col in range(1, 5): ws.cell(row=ws.max_row, column=col).font = Font(bold=True)
            
            ws.append([
                r.get("folio_mercantil", "N/A"),
                r.get("date", "N/A"),
                r.get("state", "N/A"),
                r.get("business_name", "N/A")
            ])
            ws.append([])
            
            # 2. Accionistas
            socios = r.get("socios", [])
            if socios:
                ws.append(["ACCIONISTAS / SOCIOS"])
                ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=4)
                aplicar_estilo_subheader(ws, ws.max_row, 1, 4)
                
                ws.append(["Nombre / Razón Social", "Estatus", "Acciones", "Valor de Aportación"])
                for col in range(1, 5): ws.cell(row=ws.max_row, column=col).font = Font(bold=True)
                
                for s in socios:
                    ws.append([s.get("name"), s.get("status"), s.get("shares"), s.get("value", 0.0)])
                    ws.cell(row=ws.max_row, column=4).number_format = CURRENCY_FORMAT
            else:
                ws.append(["No se encontraron accionistas registrados."])
            ws.append([])
            
            # 3. Actos y Asambleas
            actos = r.get("actos", [])
            if actos:
                ws.append(["ACTOS Y ASAMBLEAS"])
                ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=4)
                aplicar_estilo_subheader(ws, ws.max_row, 1, 4)
                
                ws.append(["Fecha", "Descripción / Acto", "No. de Documento", ""])
                for col in range(1, 4): ws.cell(row=ws.max_row, column=col).font = Font(bold=True)
                
                for a in actos:
                    ws.append([a.get("date"), a.get("description"), a.get("document_number"), ""])
                    ws.cell(row=ws.max_row, column=2).alignment = Alignment(wrap_text=True)
            else:
                ws.append(["No se encontraron actos o asambleas registrados."])
                
    else:
        ws.append(["Sin registros de entidades RPC detectados.", "", "", ""])
        ws.append([])
        ws.append([])
        
    ws.append([])
    ws.append([])
    
    # --- SECCIÓN RUG ---
    ws.append(["REGISTRO ÚNICO DE GARANTÍAS (RUG)"])
    fila_rug = ws.max_row
    ws.merge_cells(f'A{fila_rug}:G{fila_rug}')
    aplicar_estilo_header(ws, fila_rug, 1, 7)
    
    if rug_records:
        ws.append(["No. Garantía", "Acreedor", "Fecha Creación", "Vigencia", "Monto", "Moneda", "Estatus"])
        aplicar_estilo_subheader(ws, ws.max_row, 1, 7)
        for r in rug_records:
            ws.append([
                r.get("guarantee_number", "N/A"),
                r.get("creditor", "N/A"),
                r.get("creation_date", "N/A"),
                r.get("validity_date", "N/A"),
                r.get("amount", 0.0),
                r.get("currency", "MXN"),
                r.get("status", "N/A")
            ])
            # Dar formato de moneda a la columna de Monto (Columna E o 5)
            ws.cell(row=ws.max_row, column=5).number_format = CURRENCY_FORMAT
    else:
        ws.append(["Sin operaciones de garantías RUG detectadas.", "", "", "", "", "", ""])
        
    # Ajustar anchos de columna para que se lea perfecto
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 20