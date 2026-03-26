# Fluxo_IA_visual/services/report_generator/sheets/s01_executive_summary.py
from openpyxl.styles import Font, Alignment
from datetime import datetime
import re
from ..styles import (
    HEADER_FONT, HEADER_FILL, SUB_HEADER_FILL, ALERT_FILL, 
    CURRENCY_FORMAT, PERCENT_FORMAT, 
    aplicar_estilo_header, aplicar_estilo_subheader
)

def build(ws, data: dict):
    # 1. DATOS DEL CONTRIBUYENTE
    ws.append(["DATOS DEL CONTRIBUYENTE"])
    ws.merge_cells('A1:D1')
    aplicar_estilo_header(ws, 1, 1, 4)
    
    antiguedad_str = "N/A"
    reg_date_str = data.get("tax_registration_date")
    if reg_date_str:
        try:
            reg_dt = datetime.fromisoformat(reg_date_str.replace("Z", ""))
            years = (datetime.now() - reg_dt).days // 365
            antiguedad_str = f"{years} Años ({reg_date_str[:10]})"
        except: pass

    ws.append(["Razón Social", data.get("business_name", "N/A"), "Antigüedad SAT", antiguedad_str])
    ws.append(["RFC", data.get("rfc", "N/A"), "", ""])

    # 2. ESTATUS DE CREDENCIALES
    ws.append([])
    ws.append(["ESTATUS DE CREDENCIALES"])
    ws.merge_cells('A5:D5')
    aplicar_estilo_header(ws, 5, 1, 4)
    
    ws.append(["Credencial", "Estatus", "Detalle / Fecha", "Valor / Score"])
    aplicar_estilo_header(ws, 6, 1, 4)

    def format_ciec_date(date_str):
        if not date_str: return ""
        try:
            d = datetime.fromisoformat(str(date_str).replace("Z", "").split(".")[0])
            return f"{d.month}/{d.day}/{d.year}"
        except: return date_str
    
    ciec = data.get("ciec_info", {})
    buro = data.get("buro_info", {})
    opinion = data.get("compliance_opinion", {})

    ws.append(["CIEC", ciec.get("status"), "Última Rev:", ciec.get("last_check_date")])
    ws.append(["CIEC última extracción", ciec.get("status"), "última extracción", format_ciec_date(ciec.get("last_extraction_date"))])
    ws.append(["Opinión Cumpl.", opinion.get("status"), "Última Rev:", opinion.get("last_check_date")])
    ws.append(["Buró de Crédito", buro.get("status"), "Score:", buro.get("score")])

    # 3. INDICADORES DE RIESGO
    ws.append([])
    ws.append(["INDICADORES DE RIESGO"])
    fila_actual = ws.max_row
    ws.merge_cells(f'A{fila_actual}:D{fila_actual}')
    aplicar_estilo_header(ws, fila_actual, 1, 4)
    ws.append(["Indicador", "Valor", "Riesgoso", ""])
    
    risks = data.get("risk_indicators", [])
    if risks and isinstance(risks, list):
        risk_obj = risks[0]
        for k, v in risk_obj.items():
            if isinstance(v, dict):
                es_riesgoso = "SÍ" if v.get("risky") else "NO"
                ws.append([k, str(v.get("value")), es_riesgoso, ""])
                if v.get("risky"):
                    ws.cell(row=ws.max_row, column=3).fill = ALERT_FILL
    
    # 3.5 CONTRAPARTES EN LISTA NEGRA (69-B SAT)
    blacklist = data.get("blacklisted_counterparties", [])
    if blacklist:
        ws.append([])
        ws.append(["CONTRAPARTES EN LISTA NEGRA (69-B SAT)"])
        fila_bl = ws.max_row
        ws.merge_cells(f'A{fila_bl}:E{fila_bl}')
        aplicar_estilo_header(ws, fila_bl, 1, 5)
        
        ws.append(["RFC", "Razón Social", "Estatus SAT", "Facturas (Emit. / Recib.)", "Monto Total Involucrado"])
        aplicar_estilo_header(ws, ws.max_row, 1, 5)
        
        # Traductor de estatus para que el analista lo lea en español
        status_map = {
            "presumed": "Presunto",
            "dismissed": "Desvirtuado",
            "definitive": "Definitivo",
            "favorable": "Favorable"
        }
        
        for b in blacklist:
            b_rfc = b.get("rfc", "N/A")
            b_name = b.get("name", "N/A")
            b_status_raw = b.get("status", "N/A")
            b_status_es = status_map.get(b_status_raw, b_status_raw)
            
            emitidas = b.get("issued_count", 0)
            recibidas = b.get("received_count", 0)
            apariciones = f"{emitidas} / {recibidas}"
            
            # Sumamos lo que le compramos y lo que le vendimos para ver el riesgo total
            monto_total = float(b.get("issued_amount", 0.0)) + float(b.get("received_amount", 0.0))
            
            ws.append([b_rfc, b_name, b_status_es, apariciones, monto_total])
            
            # Formato de Moneda
            ws.cell(row=ws.max_row, column=5).number_format = CURRENCY_FORMAT
            
            # Alerta visual en rojo
            if b_status_raw in ["definitive", "presumed"]:
                ws.cell(row=ws.max_row, column=3).fill = ALERT_FILL
                
        # Aseguramos que la columna nueva tenga buen ancho
        ws.column_dimensions['E'].width = 25

    # 4. ACTIVIDADES ECONÓMICAS
    ws.append([])
    ws.append(["ACTIVIDADES ECONÓMICAS"])
    ws.merge_cells(f'A{ws.max_row}:D{ws.max_row}')
    aplicar_estilo_header(ws, ws.max_row, 1, 4)
    ws.append(["Actividad", "Porcentaje", "Fecha Inicio", ""])
    
    acts = data.get("economic_activities", [])
    
    # Regex para detectar caracteres de control ilegales en XML/Excel
    ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

    for act in acts:
        pct_str = f"{act.get('percentage', 0)}%"
        
        # 1. Obtenemos el texto crudo
        raw_name = str(act.get("name", "N/A"))
        
        # 2. Eliminamos caracteres ilegales
        clean_name = ILLEGAL_CHARACTERS_RE.sub("", raw_name)
        
        # 3. Quitamos saltos de línea y espacios múltiples para que se vea limpio
        clean_name = " ".join(clean_name.split())
        
        ws.append([clean_name, pct_str, act.get("start_date"), ""])# 4. ACTIVIDADES ECONÓMICAS
    ws.append([])
    ws.append(["ACTIVIDADES ECONÓMICAS"])
    ws.merge_cells(f'A{ws.max_row}:D{ws.max_row}')
    aplicar_estilo_header(ws, ws.max_row, 1, 4)
    ws.append(["Actividad", "Porcentaje", "Fecha Inicio", ""])
    
    acts = data.get("economic_activities", [])
    
    # Regex para detectar caracteres de control ilegales en XML/Excel
    ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

    for act in acts:
        pct_str = f"{act.get('percentage', 0)}%"
        
        # 1. Obtenemos el texto crudo
        raw_name = str(act.get("name", "N/A"))
        
        # 2. Eliminamos caracteres ilegales
        clean_name = ILLEGAL_CHARACTERS_RE.sub("", raw_name)
        
        # 3. Quitamos saltos de línea y espacios múltiples para que se vea limpio
        clean_name = " ".join(clean_name.split())
        
        ws.append([clean_name, pct_str, act.get("start_date"), ""])

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20

    # 4.5 EVOLUCIÓN DE EMPLEADOS (NUEVO RECUADRO)
    ws.append([])
    ws.append(["EVOLUCIÓN DE EMPLEADOS"])
    fila_emp = ws.max_row
    ws.merge_cells(f'A{fila_emp}:D{fila_emp}')
    aplicar_estilo_header(ws, fila_emp, 1, 4)
    
    ws.append(["Periodo", "Total Empleados", "Tendencia (vs Anterior)", "Diferencia"])
    aplicar_estilo_header(ws, ws.max_row, 1, 4)
    
    emp_metrics = data.get("employee_metrics")
    if emp_metrics:
        def print_emp_row(label, period_obj):
            if not period_obj:
                ws.append([label, "Sin datos", "-", "-"])
                return
            
            # Ahora accedemos como diccionario usando .get()
            total = period_obj.get("total", 0)
            trend_text = period_obj.get("trend_text", "Sin datos")
            difference = period_obj.get("difference", 0)
            
            diff_str = f"+{difference}" if difference > 0 else str(difference)
            
            ws.append([
                label, 
                total, 
                trend_text, 
                diff_str if trend_text != "Sin datos previos" else "-"
            ])

        # Extraemos usando las llaves del diccionario
        print_emp_row("Hace 24 Meses", emp_metrics.get("month_24"))
        print_emp_row("Hace 12 Meses", emp_metrics.get("month_12"))
        print_emp_row("Hace 9 Meses",  emp_metrics.get("month_9"))
        print_emp_row("Hace 6 Meses",  emp_metrics.get("month_6"))
        print_emp_row("Hace 3 Meses",  emp_metrics.get("month_3"))
        print_emp_row("Mes Actual",    emp_metrics.get("month_1"))
    else:
        ws.append(["Sin historial de empleados disponible.", "", "", ""])
        ws.merge_cells(f'A{ws.max_row}:D{ws.max_row}')

    # 5. RESUMEN DE BURÓ
    summary = buro.get("summary_metrics")
    if summary:
        ws.append([])
        ws.append(["RESUMEN DE BURÓ"])
        ws.merge_cells(f'A{ws.max_row}:D{ws.max_row}')
        aplicar_estilo_header(ws, ws.max_row, 1, 4)

        trend_pct = summary.get("inquiries_trend_pct", 0.0)
        if trend_pct > 0:
            ws.append(["", "Han ido aumentando sus consultas", trend_pct, ""])
            ws.cell(row=ws.max_row, column=3).number_format = PERCENT_FORMAT
        elif trend_pct < 0:
            ws.append(["", "Han ido disminuyendo sus consultas", trend_pct, ""])
            ws.cell(row=ws.max_row, column=3).number_format = PERCENT_FORMAT
        else:
            ws.append(["", "Estables", "0%", ""])
            
        ws.append([])
        
        metricas_globales = [
            ("Monto Máximo Abierto", summary.get("total_open_max_amount", 0.0), CURRENCY_FORMAT),
            ("Saldo Vigente", summary.get("total_current_balance", 0.0), CURRENCY_FORMAT),
            ("Saldo Vencido Total", summary.get("total_past_due", 0.0), CURRENCY_FORMAT),
            ("Saldo Vencido Abierto", summary.get("total_open_past_due", 0.0), CURRENCY_FORMAT),
            ("Pago Mensual 1", summary.get("monthly_payment_1", 0.0), CURRENCY_FORMAT),
            ("Pago Mensual 2", summary.get("monthly_payment_2", 0.0), CURRENCY_FORMAT),
            ("Plazo Ponderado (años)", summary.get("weighted_term_years", 0.0), '0.00')
        ]
        for titulo, valor, fmt in metricas_globales:
            ws.append([titulo, valor, "", ""])
            ws.cell(row=ws.max_row, column=2).number_format = fmt

        ws.append([])
        
        def print_bucket(label, key):
            b_data = summary.get(key, {})
            amt = b_data.get("amount", 0.0)
            pct = b_data.get("percentage", 0.0)
            val_str = amt if amt > 0 else "-"
            ws.append([label, val_str, pct if amt > 0 else "", ""])
            if amt > 0:
                ws.cell(row=ws.max_row, column=2).number_format = CURRENCY_FORMAT
                ws.cell(row=ws.max_row, column=3).number_format = PERCENT_FORMAT
            else:
                ws.cell(row=ws.max_row, column=2).alignment = Alignment(horizontal='right')

        print_bucket("saldoVencidoDe1a29Dias", "bucket_1_29")
        print_bucket("saldoVencidoDe30a59Dias", "bucket_30_59")
        print_bucket("saldoVencidoDe60a89Dias", "bucket_60_89")
        print_bucket("saldoVencidoDe90a119Dias", "bucket_90_119")
        print_bucket("saldoVencidoDe120a179Dias", "bucket_120_179")
        print_bucket("saldoVencidoDe180DiasOMas", "bucket_180_plus")

    # 6. RESUMEN FINANCIERO (BALANCE Y RESULTADOS)
    financials = data.get("financial_ratios_history", [])
    if financials:
        financials_sorted = sorted(financials, key=lambda x: str(x.get("year", "")))
        years = [str(f.get("year", "N/A")) for f in financials_sorted]
        max_cols = len(years) + 1

        ws.append([])
        ws.append(["RESUMEN FINANCIERO (BALANCE Y RESULTADOS)"] + [""] * (len(years) - 1))
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=max_cols)
        aplicar_estilo_header(ws, ws.max_row, 1, max_cols)

        ws.append(["Concepto"] + years)
        aplicar_estilo_subheader(ws, ws.max_row, 1, max_cols)

        def print_fin_row(label, key, is_bold=False):
            row_data = [label]
            for f in financials_sorted:
                row_data.append(f.get(key, 0.0))
            ws.append(row_data)
            
            curr_row = ws.max_row
            if is_bold: ws.cell(row=curr_row, column=1).font = Font(bold=True)
            for col_idx in range(2, max_cols + 1):
                ws.cell(row=curr_row, column=col_idx).number_format = CURRENCY_FORMAT

        ws.append(["BALANCE GENERAL"] + [""] * len(years))
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, italic=True)
        
        print_fin_row("Activo", "input_assets")
        print_fin_row("  Activo Corto Plazo", "input_assets_short_term")
        print_fin_row("  Activo Largo Plazo", "input_assets_long_term")
        
        print_fin_row("Pasivo", "input_liabilities")
        print_fin_row("  Pasivo Corto Plazo", "input_liabilities_short_term")
        print_fin_row("  Pasivo Largo Plazo", "input_liabilities_long_term")
        
        print_fin_row("Capital Contable", "input_equity")
        print_fin_row("  Capital Social", "input_equity_social")
        
        ws.append([])

        ws.append(["ESTADO DE RESULTADOS"] + [""] * len(years))
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, italic=True)

        print_fin_row("Ingresos Netos", "input_revenue")
        print_fin_row("Utilidad (Pérdida) Bruta", "input_gross_profit")
        print_fin_row("Utilidad (Pérdida) de Operación", "ebit")
        print_fin_row("Utilidad (Pérdida) antes de Impuestos", "ebt")
        print_fin_row("Impuestos a la Utilidad", "input_taxes")
        print_fin_row("Utilidad (Pérdida) Neta", "input_net_income", is_bold=True)