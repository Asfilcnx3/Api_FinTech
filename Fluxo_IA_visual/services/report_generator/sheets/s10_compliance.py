# Fluxo_IA_visual/services/report_generator/sheets/s10_compliance.py
from openpyxl.styles import Alignment, Font
from ..styles import aplicar_estilo_header, aplicar_estilo_subheader

def build(ws, data: dict):
    ws.append(["ANÁLISIS DE LA OPINIÓN DE CUMPLIMIENTO (32-D) CON IA"])
    ws.merge_cells('A1:C1')
    aplicar_estilo_header(ws, 1, 1, 3)
    
    ws.append([])
    
    # Extraemos el nuevo objeto estructurado
    compliance_data = data.get("compliance_llm_data")
    
    if not compliance_data:
        ws.append(["Análisis no disponible o PDF no procesado."])
        return

    # --- 1. OPINIÓN GENERAL DE LA IA ---
    ws.append(["Conclusión del Auditor Virtual (IA):"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    
    # Extraemos de forma segura ya sea que llegue como dict o como objeto Pydantic
    if isinstance(compliance_data, dict):
        opinion = compliance_data.get("opinion_ia", "Sin opinión.")
        obligaciones = compliance_data.get("obligaciones_omitidas", [])
    else:
        opinion = getattr(compliance_data, "opinion_ia", "Sin opinión.")
        obligaciones = getattr(compliance_data, "obligaciones_omitidas", [])
    
    ws.append([opinion])
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=3)
    ws.cell(row=ws.max_row, column=1).alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[ws.max_row].height = 45
    
    ws.append([])

    # --- 2. TABLA DE OBLIGACIONES OMITIDAS ---
    if obligaciones:
        ws.append(["OBLIGACIONES FISCALES OMITIDAS (ALERTA)"])
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=2)
        aplicar_estilo_subheader(ws, ws.max_row, 1, 2)
        
        ws.append(["Impuesto / Obligación Omitida", "Periodo(s)"])
        for col in range(1, 3):
            ws.cell(row=ws.max_row, column=col).font = Font(bold=True)
            ws.cell(row=ws.max_row, column=col).alignment = Alignment(horizontal='center')
            
        for obs in obligaciones:
            # Extraemos de forma segura los atributos de la lista interna
            if isinstance(obs, dict):
                impuesto = obs.get("impuesto", "N/A")
                periodos = obs.get("periodos", "N/A")
            else:
                impuesto = getattr(obs, "impuesto", "N/A")
                periodos = getattr(obs, "periodos", "N/A")
                
            ws.append([impuesto, periodos])
            ws.cell(row=ws.max_row, column=1).alignment = Alignment(wrap_text=True)
            ws.cell(row=ws.max_row, column=2).alignment = Alignment(horizontal='center', wrap_text=True)
            
    else:
        # Si la lista viene vacía, le damos la buena noticia al analista
        ws.append(["ESTATUS POSITIVO: No se detectaron obligaciones fiscales omitidas en el reporte."])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, color="008000") # Verde
        
    # Ajuste de anchos para que se lea perfecto
    ws.column_dimensions['A'].width = 55
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 20