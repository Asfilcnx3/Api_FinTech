# Fluxo_IA_visual/services/report_generator/sheets/s11_concentration_history.py
from openpyxl.styles import Alignment
from ..styles import (
    HEADER_FONT, SUB_HEADER_FILL, 
    CURRENCY_FORMAT, aplicar_estilo_header
)

def build(ws, data: dict):
    concentration = data.get("concentration_last_12m", {})
    top_clients = concentration.get("top_5_clients", [])
    top_suppliers = concentration.get("top_5_suppliers", [])

    # 1. Recolectar todas las fechas únicas para hacer las columnas dinámicas
    all_dates = set()
    for item in top_clients + top_suppliers:
        for t in item.get("monthly_history", []):
            if t.get("date"):
                all_dates.add(t.get("date"))
                
    if not all_dates:
        ws.append(["No hay historial mensual de transacciones disponible."])
        return

    sorted_dates = sorted(list(all_dates)) # Orden cronológico (ej. 2025-03, 2025-04...)

    # 2. Configurar Encabezados
    base_headers = ["Tipo", "Nombre", "RFC", "Monto Total"]
    full_headers = base_headers + sorted_dates
    
    ws.append(["HISTORIAL MENSUAL: TOP 5 CLIENTES Y PROVEEDORES"])
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=len(full_headers))
    aplicar_estilo_header(ws, ws.max_row, 1, len(full_headers))

    ws.append(full_headers)
    sub_row = ws.max_row
    for col in range(1, len(full_headers) + 1):
        cell = ws.cell(row=sub_row, column=col)
        cell.font = HEADER_FONT
        cell.fill = SUB_HEADER_FILL
        cell.alignment = Alignment(horizontal='center')

    # 3. Función para rellenar la matriz
    def pintar_matriz(tipo_label, lista_items):
        for item in lista_items:
            row_data = [
                tipo_label,
                item.get("name", "N/A"),
                item.get("rfc", "N/A"),
                item.get("total_amount", 0.0)
            ]
            
            # Convertimos el historial a un diccionario rápido { "2025-03": 1500.0, ... }
            historial_dict = {t.get("date"): t.get("total", 0.0) for t in item.get("monthly_history", [])}
            
            # Rellenar los meses (0.0 si no facturó ese mes)
            for d in sorted_dates:
                row_data.append(historial_dict.get(d, 0.0))
                
            ws.append(row_data)
            
            # Aplicar formato de moneda
            curr_row = ws.max_row
            ws.cell(row=curr_row, column=1).alignment = Alignment(horizontal='center') # Tipo
            for col_idx in range(4, len(full_headers) + 1): # Desde "Monto Total" hasta el final
                ws.cell(row=curr_row, column=col_idx).number_format = CURRENCY_FORMAT

    # 4. Inyectar datos
    pintar_matriz("CLIENTE", top_clients)
    ws.append([]) # Fila en blanco separadora
    pintar_matriz("PROVEEDOR", top_suppliers)

    # 5. Anchos de Columna
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    
    # Columnas de fechas
    for i in range(4, len(full_headers)):
        ws.column_dimensions[chr(64 + i + 1)].width = 18 # A partir de la E