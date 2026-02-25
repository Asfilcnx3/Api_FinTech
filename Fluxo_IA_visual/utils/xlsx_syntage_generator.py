import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle, Border, Side
from datetime import datetime

def generar_excel_syntage(data: dict) -> bytes:
    wb = Workbook()
    
    # --- ESTILOS ---
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    sub_header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    currency_style = NamedStyle(name='currency_style', number_format='$#,##0.00')
    percent_style = NamedStyle(name='percent_style', number_format='0.00%')
    date_style = NamedStyle(name='date_style', number_format='YYYY-MM-DD')
    alert_fill = PatternFill(start_color="C0504D", end_color="C0504D", fill_type="solid")

    def estilo_header(ws, row_idx, col_start=1, col_end=None):
        if col_end is None: col_end = ws.max_column
        for col in range(col_start, col_end + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

    # ==========================================
    # HOJA 1: RESUMEN EJECUTIVO (MODIFICADA)
    # ==========================================
    ws1 = wb.active
    ws1.title = "Resumen Ejecutivo"
    
    # 1. Datos Contribuyente (+ Antigüedad)
    ws1.append(["DATOS DEL CONTRIBUYENTE"])
    ws1.merge_cells('A1:D1') # Expandimos a 4 columnas por el requerimiento de dividir celdas
    estilo_header(ws1, 1, 1, 4)
    
    # Cálculo Antigüedad
    antiguedad_str = "N/A"
    reg_date_str = data.get("tax_registration_date")
    if reg_date_str:
        try:
            reg_dt = datetime.fromisoformat(reg_date_str.replace("Z", ""))
            years = (datetime.now() - reg_dt).days // 365
            antiguedad_str = f"{years} Años ({reg_date_str[:10]})"
        except: pass

    ws1.append(["Razón Social", data.get("business_name", "N/A"), "Antigüedad SAT", antiguedad_str])
    ws1.append(["RFC", data.get("rfc", "N/A"), "", ""]) # Merge visual opcional

    # 2. Credenciales (CELDA PARTIDA SOLICITADA)
    ws1.append([])
    ws1.append(["ESTATUS DE CREDENCIALES"])
    ws1.merge_cells('A5:D5')
    estilo_header(ws1, 5, 1, 4)
    
    ciec = data.get("ciec_info", {})
    buro = data.get("buro_info", {})
    opinion = data.get("compliance_opinion", {})

    # Formato solicitado: Col A: Nombre, Col B: Estatus, Col C: Label Extra, Col D: Valor Extra
    ws1.append(["Credencial", "Estatus", "Detalle / Fecha", "Valor / Score"])
    estilo_header(ws1, 6, 1, 4)
    
    # Fila CIEC
    ws1.append(["CIEC", ciec.get("status"), "Última Rev:", ciec.get("last_check_date")])
    # Fila Opinión
    ws1.append(["Opinión Cumpl.", opinion.get("status"), "Última Rev:", opinion.get("last_check_date")])
    # Fila Buró (Split Solicitado)
    ws1.append(["Buró de Crédito", buro.get("status"), "Score:", buro.get("score")])

    # 3. Riesgos
    ws1.append([])
    ws1.append(["INDICADORES DE RIESGO"])
    ws1.merge_cells('A11:D11')
    estilo_header(ws1, 11, 1, 4)
    ws1.append(["Indicador", "Valor", "Riesgoso", ""])
    
    risks = data.get("risk_indicators", [])
    if risks and isinstance(risks, list):
        risk_obj = risks[0]
        for k, v in risk_obj.items():
            if isinstance(v, dict):
                es_riesgoso = "SÍ" if v.get("risky") else "NO"
                ws1.append([k, str(v.get("value")), es_riesgoso, ""])
                if v.get("risky"):
                    ws1.cell(row=ws1.max_row, column=3).fill = alert_fill

    # 4. Actividades Económicas (NUEVO)
    ws1.append([])
    ws1.append(["ACTIVIDADES ECONÓMICAS"])
    ws1.merge_cells(f'A{ws1.max_row}:D{ws1.max_row}')
    estilo_header(ws1, ws1.max_row, 1, 4)
    
    ws1.append(["Actividad", "Porcentaje", "Fecha Inicio", ""])
    acts = data.get("economic_activities", [])
    for act in acts:
        # Formatear porcentaje
        pct = act.get("percentage", 0)
        pct_str = f"{pct}%"
        ws1.append([act.get("name"), pct_str, act.get("start_date"), ""])

    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 20
    ws1.column_dimensions['C'].width = 20
    ws1.column_dimensions['D'].width = 20

    # ==========================================
    # HOJA 2: FACTURACIÓN EN EL TIEMPO (RENOVADA)
    # ==========================================
    ws2 = wb.create_sheet("Facturación en el tiempo")
    
    # Matriz solicitada:
    # Cols: Last 24, 12, 9, 6, 3
    # Rows (Bloques): Revenue, Expenses, Inflows, Outflows, NFCF
    # Sub-Rows: Mean, Median Growth, Slope, CAGR
    
    periods_order = ["last_24_months", "last_12_months", "last_9_months", "last_6_months", "last_3_months"]
    metrics_map = ["revenue", "expenditures", "inflows", "outflows", "nfcf"]
    sub_metrics = [
        ("Promedio (Mean)", "mean", currency_style), # Media pura (moneda)
        ("Mediana (Median)", "median", currency_style), # Mediana pura (moneda)
        ("Crecimiento vs Periodo Anterior", "period_growth_rate", percent_style), # Crecimiento real (%)
        ("Slope (Tendencia Lineal)", "linear_slope", currency_style), # Pendiente de la tendencia (moneda)
        ("CAGR / CMGR", "cagr_cmgr", percent_style) # Crecimiento Compuesto (%)
    ]
    
    # Encabezados de Columnas
    ws2.append(["Métrica / Periodo"] + [p.replace("_", " ").title() for p in periods_order])
    estilo_header(ws2, 1)
    
    stats = data.get("stats_last_months", {})

    for metric_key in metrics_map:
        # Título de Sección (ej. REVENUE)
        ws2.append([metric_key.upper()])
        # Pintar fila de título sección
        curr_row = ws2.max_row
        ws2.cell(row=curr_row, column=1).fill = sub_header_fill
        ws2.cell(row=curr_row, column=1).font = header_font
        
        # Generar las 4 filas de sub-métricas
        for label, sub_key, style in sub_metrics:
            row_data = [label]
            for period in periods_order:
                # Navegar: stats -> period -> metric -> sub_metric
                # ej: stats['last_3_months']['revenue']['mean']
                p_data = stats.get(period) or {}
                m_data = p_data.get(metric_key) or {}
                val = m_data.get(sub_key, 0)
                row_data.append(val)
            
            ws2.append(row_data)
            
            # Aplicar estilos a las celdas de datos
            curr_row = ws2.max_row
            for col_idx in range(2, len(periods_order) + 2):
                cell = ws2.cell(row=curr_row, column=col_idx)
                cell.style = style

        # Espacio vacío
        ws2.append([""])

    ws2.column_dimensions['A'].width = 35
    for i in range(2, 7):
        ws2.column_dimensions[chr(64+i)].width = 20

    # ==========================================
    # HOJA 3: PROYECCIONES (ACTUALIZADA)
    # ==========================================
    ws3 = wb.create_sheet("Proyecciones (Escenarios)")
    
    predictions = data.get("financial_predictions", {})
    metrics_to_export = ["revenue", "expenditures", "inflows", "outflows", "nfcf"]
    raw_history = data.get("raw_data_history", [])
    
    # 1. Obtener Fechas: 12 Pasadas + 12 Futuras
    # Extraemos fechas históricas (últimos 12 meses de la raw data)
    hist_dates_iso = [x["date"] for x in raw_history[-12:]] if raw_history else []
    
    # Extraemos fechas futuras del primer modelo
    future_dates_iso = []
    first_metric = predictions.get("revenue", {})
    if first_metric and first_metric.get("linear"):
        future_dates_iso = [pt["date"] for pt in first_metric["linear"]["scenarios"]["realistic"]]
    
    # Encabezados
    headers = ["Métrica", "Modelo", "Escenario", "Growth Comp.", "Growth Proy."] 
    # Añadir fechas históricas (Colapsadas visualmente o marcadas distinto)
    headers += [f"Hist: {d}" for d in hist_dates_iso]
    # Añadir fechas futuras
    headers += [f"Proy: {d}" for d in future_dates_iso]
    
    ws3.append(headers)
    estilo_header(ws3, 1)

    for metric in metrics_to_export:
        m_data = predictions.get(metric, {})
        if not m_data: continue
        
        # Obtener valores históricos para esta métrica (últimos 12)
        # Mapeamos nombre métrica a clave en raw_data_history
        metric_key_map = {
            "revenue": "revenue", "expenditures": "expenses",
            "inflows": "inflows_amount", "outflows": "outflows_amount",
            "nfcf": "nfcf"
        }
        key = metric_key_map.get(metric)
        hist_vals = [item.get(key, 0.0) for item in raw_history[-12:]]
        
        # Rellenar ceros si falta historia
        if len(hist_vals) < 12:
            hist_vals = [0.0]*(12-len(hist_vals)) + hist_vals

        for model_type in ["linear", "exponential", "seasonal"]:
            model_res = m_data.get(model_type)
            if not model_res: continue
            
            scenarios = model_res.get("scenarios", {})
            
            # Métricas de crecimiento
            # Comparison: Forecast Total vs History Total
            g_comp_r = model_res.get('growth_realistic') 
            g_comp_o = model_res.get('growth_optimistic')
            g_comp_p = model_res.get('growth_pessimistic')

            # Projection: Trend interna (campo nuevo agregado en Forecaster)
            g_proy_r = model_res.get('trend_realistic') 
            g_proy_o = model_res.get('trend_optimistic')
            g_proy_p = model_res.get('trend_pessimistic')
            
            # Función helper para formatear fila
            def make_row(label, comp, proy, futures):
                row = [
                    metric.upper() if label == "Realista" else "", # Métrica
                    model_type.title() if label == "Realista" else "", # Modelo
                    label, # Escenario
                    comp, # Growth Comparison
                    proy  # Growth Projection
                ]
                # Agregamos historia (Siempre la REAL para contexto visual)
                row += hist_vals 
                # Agregamos futuro
                row += [pt.get("value") for pt in futures]
                return row

            # Fila Realista
            ws3.append(make_row("Realista", g_comp_r, g_proy_r, scenarios.get("realistic", [])))
            # Fila Optimista
            ws3.append(make_row("Optimista (+)", g_comp_o, g_proy_o, scenarios.get("optimistic", [])))
            # Fila Pesimista
            ws3.append(make_row("Pesimista (-)", g_comp_p, g_proy_p, scenarios.get("pessimistic", [])))
            
            ws3.append([""]) # Separador visual

    # Estilos Proyecciones
    ws3.column_dimensions['A'].width = 15
    ws3.column_dimensions['B'].width = 15
    ws3.column_dimensions['C'].width = 20
    ws3.column_dimensions['D'].width = 15
    ws3.column_dimensions['E'].width = 15
    
    # Formato Porcentaje (Cols D y E)
    for row in ws3.iter_rows(min_row=2, min_col=4, max_col=5):
        for cell in row: cell.style = percent_style
        
    # Formato Moneda (Desde Col F hasta el final)
    max_col = ws3.max_column
    for row in ws3.iter_rows(min_row=2, min_col=6, max_col=max_col):
        for cell in row: 
            if isinstance(cell.value, (int, float)):
                cell.style = currency_style

    # Separador visual entre historia y futuro
    # La columna donde empieza el futuro es 6 (A-E) + 12 (hist) = 18?
    # A=1, B=2, C=3, D=4, E=5. Historia es 6 a 17. Futuro empieza en 18.
    col_split = 5 + len(hist_vals)
    for row in range(1, ws3.max_row + 1):
        cell = ws3.cell(row=row, column=col_split)
        cell.border = Border(right=Side(style='medium', color="000000"))


    # ==========================================
    # HOJA 4: RAW DATA (NUEVA PAGINA)
    # ==========================================
    ws_raw = wb.create_sheet("Raw Data")
    
    # Encabezados solicitados
    headers_raw = [
        "startDate", "Revenue", "Expenses", 
        "Inflows mxn amounts", "Outflows mxn amounts", 
        "NFCF", "Inflows transactions", "Outflows transactions"
    ]
    ws_raw.append(headers_raw)
    estilo_header(ws_raw, 1)
    
    for item in raw_history:
        # Item es un dict (pydantic model dump)
        ws_raw.append([
            item.get("date"),
            item.get("revenue"),
            item.get("expenses"),
            item.get("inflows_amount"),
            item.get("outflows_amount"),
            item.get("nfcf"),
            item.get("inflows_transactions"),
            item.get("outflows_transactions")
        ])

    # Formatos Raw Data
    ws_raw.column_dimensions['A'].width = 15
    for i in range(2, 9): ws_raw.column_dimensions[chr(64+i)].width = 20
    
    # Moneda para cols B a F
    for row in ws_raw.iter_rows(min_row=2, min_col=2, max_col=6):
        for cell in row: cell.style = currency_style

    # ==========================================
    # HOJA 5: ESTADOS FINANCIEROS Y ÁRBOL
    # ==========================================
    ws_fin = wb.create_sheet("Estados Financieros")
    
    financials = data.get("financial_ratios_history", [])
    
    if not financials:
        ws_fin.append(["No se encontró información de estados financieros."])
    else:
        # Aseguramos que los años estén ordenados de más reciente a más antiguo
        financials_sorted = sorted(financials, key=lambda x: str(x.get("year", "")), reverse=True)
        years = [str(f.get("year", "N/A")) for f in financials_sorted]
        
        # 1. Encabezados (Años)
        ws_fin.append(["Concepto / Año"] + years)
        estilo_header(ws_fin, 1, 1, len(years) + 1)
        
        # 2. Mapeo de Conceptos Financieros (Crudos y Razones)
        conceptos_financieros = [
            ("--- DATOS CONSOLIDADOS ---", None),
            ("Activos Totales", "input_assets"),
            ("Capital Contable (Equity)", "input_equity"),
            ("Ingresos Netos (Revenue)", "input_revenue"),
            ("Impuestos Calculados", "input_taxes"),
            ("Utilidad Neta Consolidada", "input_net_income"),
            ("EBIT Consolidado", "ebit"),
            ("EBT Consolidado", "ebt"),
            ("NOPAT", "nopat"),
            
            ("", None), # Fila en blanco
            ("--- DATOS CRUDOS EXTRAÍDOS ---", None),
            ("Utilidad Neta (Crudo)", "raw_net_profit"),
            ("Pérdida Neta (Crudo)", "raw_net_loss"),
            ("Utilidad Operativa (Crudo)", "raw_ebit_profit"),
            ("Pérdida Operativa (Crudo)", "raw_ebit_loss"),
            ("Utilidad antes de Imp. (Crudo)", "raw_ebt_profit"),
            ("Pérdida antes de Imp. (Crudo)", "raw_ebt_loss"),
            
            ("", None), # Fila en blanco
            ("--- RAZONES FINANCIERAS ---", None),
            ("ROA", "roa"),
            ("ROE", "roe"),
            ("Margen Neto (%)", "net_profit_margin_percent")
        ]
        
        # 3. Construcción de Filas Iniciales
        for label, key in conceptos_financieros:
            row_data = [label]
            if key is None:
                for _ in financials_sorted: row_data.append("")
                ws_fin.append(row_data)
                ws_fin.cell(row=ws_fin.max_row, column=1).font = Font(bold=True)
                continue
                
            for f_year in financials_sorted:
                val = f_year.get(key, 0.0)
                row_data.append(val)
            ws_fin.append(row_data)
            
            curr_row = ws_fin.max_row
            for col_idx in range(2, len(years) + 2):
                cell = ws_fin.cell(row=curr_row, column=col_idx)
                if key in ["roa", "roe", "net_profit_margin_percent"]:
                    cell.style = percent_style
                    if key != "net_profit_margin_percent": 
                        cell.value = cell.value / 100 if cell.value else 0
                else:
                    cell.style = currency_style

        # Ajuste de Anchos de Columna iniciales
        ws_fin.column_dimensions['A'].width = 40
        for i in range(2, len(years) + 2):
            ws_fin.column_dimensions[chr(64+i)].width = 20

    # ---------------------------------------------------------
    # DESGLOSE COMPLETO DEL ÁRBOL FINANCIERO (Misma Hoja)
    # ---------------------------------------------------------
    fs_tree = data.get("financial_statements_tree", {})
    
    # 1. Función recursiva TOTALMENTE CIEGA para encontrar años
    all_years_tree = set()
    def extract_years_from_tree(node):
        if isinstance(node, dict):
            for k, v in node.items():
                if str(k).isdigit() and len(str(k)) == 4:
                    all_years_tree.add(str(k))
                elif isinstance(v, (dict, list)):
                    extract_years_from_tree(v)
        elif isinstance(node, list):
            for item in node:
                extract_years_from_tree(item)

    extract_years_from_tree(fs_tree)
    sorted_tree_years = sorted(list(all_years_tree), reverse=True)

    if sorted_tree_years:
        # 2. Función recursiva que escribe celda por celda y rastrea la fila actual
        def write_tree_node(ws, node, current_row, start_col, depth=0):
            if isinstance(node, dict) and "data" in node:
                return write_tree_node(ws, node["data"], current_row, start_col, depth)
            
            if isinstance(node, list):
                for item in node:
                    current_row = write_tree_node(ws, item, current_row, start_col, depth)
                return current_row
                
            if isinstance(node, dict):
                category = node.get("category", "")
                
                if category:
                    indent = "    " * depth
                    # Escribir categoría
                    ws.cell(row=current_row, column=start_col, value=f"{indent}{category}")
                    
                    if depth == 0:
                        ws.cell(row=current_row, column=start_col).font = Font(bold=True)
                        
                    # Escribir valores de los años
                    col_offset = 1
                    for y in sorted_tree_years:
                        val_node = node.get(y, 0.0)
                        val = 0.0
                        if isinstance(val_node, dict):
                            val = float(val_node.get("Total") or val_node.get("total") or 0.0)
                        elif isinstance(val_node, (int, float)):
                            val = float(val_node)
                        
                        c = ws.cell(row=current_row, column=start_col + col_offset, value=val)
                        c.style = currency_style
                        col_offset += 1
                        
                    current_row += 1
                
                children = node.get("children", [])
                if children:
                    next_depth = depth + 1 if category else depth
                    current_row = write_tree_node(ws, children, current_row, start_col, next_depth)
            
            return current_row

        # 3. Determinar coordenadas de inicio
        start_row_trees = ws_fin.max_row + 3
        col_bs = 1  # Balance Sheet empieza en columna A (1)
        col_is = len(sorted_tree_years) + 3  # Income Statement empieza dejando 1 columna vacía
        
        # --- ENCABEZADOS BALANCE ---
        ws_fin.cell(row=start_row_trees, column=col_bs, value="ESTADO DE POSICIÓN FINANCIERA")
        ws_fin.cell(row=start_row_trees, column=col_bs).font = Font(bold=True, size=12)
        
        ws_fin.cell(row=start_row_trees + 1, column=col_bs, value="Categoría")
        for i, y in enumerate(sorted_tree_years):
            ws_fin.cell(row=start_row_trees + 1, column=col_bs + i + 1, value=y)
        estilo_header(ws_fin, start_row_trees + 1, col_bs, col_bs + len(sorted_tree_years))
        
        # --- ENCABEZADOS RESULTADOS ---
        ws_fin.cell(row=start_row_trees, column=col_is, value="ESTADO DE RESULTADOS")
        ws_fin.cell(row=start_row_trees, column=col_is).font = Font(bold=True, size=12)
        
        ws_fin.cell(row=start_row_trees + 1, column=col_is, value="Categoría")
        for i, y in enumerate(sorted_tree_years):
            ws_fin.cell(row=start_row_trees + 1, column=col_is + i + 1, value=y)
        estilo_header(ws_fin, start_row_trees + 1, col_is, col_is + len(sorted_tree_years))
        
        # Ajustar anchos de columnas para el lado derecho
        col_letter_cat = ws_fin.cell(row=1, column=col_is).column_letter
        ws_fin.column_dimensions[col_letter_cat].width = 40
        for i in range(len(sorted_tree_years)):
            col_letter_year = ws_fin.cell(row=1, column=col_is + i + 1).column_letter
            ws_fin.column_dimensions[col_letter_year].width = 20

        # --- IMPRIMIR ÁRBOLES LADO A LADO ---
        data_start_row = start_row_trees + 2
        
        bs_tree = fs_tree.get("balance_sheet", {})
        write_tree_node(ws_fin, bs_tree, current_row=data_start_row, start_col=col_bs)
        
        is_tree = fs_tree.get("income_statement", {})
        write_tree_node(ws_fin, is_tree, current_row=data_start_row, start_col=col_is)

    # ==========================================
    # HOJA 5: DETALLE DE BURÓ (EXPANDIDO)
    # ==========================================
    ws4 = wb.create_sheet("Detalle de Buró")
    
    buro_info = data.get("buro_info", {})
    
    # ---------------------------------------------------------
    # FIX: Extraemos la data correcta sin importar qué nivel de anidación tenga
    # ---------------------------------------------------------
    raw_buro_container = buro_info.get("raw_buro_data", {})
    
    # A veces Syntage envuelve la data en otra llave "data" o "respuesta"
    raw_buro = raw_buro_container
    if "data" in raw_buro_container:
        raw_buro = raw_buro_container["data"]
    elif "respuesta" in raw_buro_container:
        raw_buro = raw_buro_container["respuesta"]
        
    # --- FUNCIONES AUXILIARES PARA TABLAS DINÁMICAS ---
    def build_kv_table(ws, title, dict_data):
        if not dict_data or not isinstance(dict_data, dict): return
        ws.append([title])
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=2)
        estilo_header(ws, ws.max_row, 1, 2)
        for k, v in dict_data.items():
            if not isinstance(v, (dict, list)): # Ignoramos anidados para mantenerlo plano
                ws.append([str(k), str(v)])
        ws.append([""]) # Separador

    def build_list_table(ws, title, list_data, cols_map):
        if not list_data: return
        
        # Fix: A veces (como en historialConsultas de empresas) viene como un solo dict, no como lista
        if isinstance(list_data, dict): 
            # Si el dict tiene las llaves adentro, lo envolvemos en lista
            if cols_map[0][1] in list_data:
                list_data = [list_data]
            else:
                # Si es un dict de dicts (menos común), extraemos los valores
                list_data = list(list_data.values())
                
        if not isinstance(list_data, list): return
        
        ws.append([title])
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=len(cols_map))
        estilo_header(ws, ws.max_row, 1, len(cols_map))
        
        headers = [c[0] for c in cols_map]
        ws.append(headers)
        sub_head_row = ws.max_row
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=sub_head_row, column=col)
            cell.font = header_font
            cell.fill = sub_header_fill
            cell.alignment = Alignment(horizontal='center')
            
        for item in list_data:
            if isinstance(item, dict):
                row = [str(item.get(c[1], "")) for c in cols_map]
                ws.append(row)
        ws.append([""]) # Separador

    # ---------------------------------------------------------
    # 1. DATOS GENERALES Y ENCABEZADO
    # ---------------------------------------------------------
    build_kv_table(ws4, "ENCABEZADO / METADATOS", raw_buro.get("encabezado", {}))
    build_kv_table(ws4, "DATOS GENERALES (EMPRESA)", raw_buro.get("datosGenerales", {}))
    
    # Manejo si es Persona Física (PF)
    persona_node = raw_buro.get("persona", {})
    if not persona_node and "respuesta" in raw_buro:
        persona_node = raw_buro.get("respuesta", {}).get("persona", {})
        
    if persona_node:
        build_kv_table(ws4, "DATOS PERSONA (PF)", persona_node.get("nombre", {}))
        
        # Domicilios PF
        domicilios = persona_node.get("domicilios", {}).get("domicilio", [])
        build_list_table(ws4, "DOMICILIOS", domicilios, [
            ("Calle", "direccion1"), ("Colonia", "coloniaPoblacion"), ("Ciudad", "ciudad"), 
            ("Estado", "estado"), ("CP", "cp"), ("Fecha Registro", "fechaRegistroDomicilio")
        ])
        # Empleos PF
        empleos = persona_node.get("empleos", {}).get("empleo", [])
        build_list_table(ws4, "EMPLEOS", empleos, [
            ("Empresa", "nombreEmpresa"), ("Puesto", "puesto"), ("Salario", "salario"),
            ("Calle", "direccion1"), ("Colonia", "coloniaPoblacion"), ("Estado", "estado")
        ])

    # ---------------------------------------------------------
    # 2. SCORES DETALLADOS
    # ---------------------------------------------------------
    score_list = raw_buro.get("score", [])
    if not score_list and "scoreBuroCredito" in persona_node:
        score_list = persona_node["scoreBuroCredito"]
    build_list_table(ws4, "SCORE DE CRÉDITO DETALLADO", score_list, [
        ("Nombre Score", "nombreScore"), ("Valor Score", "valorScore"), ("Código Score", "codigoScore"), 
        ("Razón 1", "codigoRazon1"), ("Razón 2", "codigoRazon2"), ("Razón 3", "codigoRazon3"), ("Error", "errorScore")
    ])

    # ---------------------------------------------------------
    # 3. ALERTAS (HAWK)
    # ---------------------------------------------------------
    hawk_list = raw_buro.get("hawkHr", [])
    build_list_table(ws4, "ALERTAS (HAWK)", hawk_list, [
        ("Código", "codigoHawk"), ("Fecha", "fechaMensajeHawk"), 
        ("Reporta", "tipoUsuarioReporta"), ("Descripción", "descripcionPrevencionHawk")
    ])

    # ---------------------------------------------------------
    # 4. PARÁMETROS DE CALIFICACIÓN
    # ---------------------------------------------------------
    califica_list = raw_buro.get("califica", [])
    build_list_table(ws4, "PARÁMETROS DE CALIFICACIÓN", califica_list, [
        ("Clave", "clave"), ("Nombre / Métrica", "nombre"), ("Valor", "valorCaracteristica")
    ])

    # ---------------------------------------------------------
    # 5. ACCIONISTAS (Empresas)
    # ---------------------------------------------------------
    accionistas_list = raw_buro.get("accionista", [])
    build_list_table(ws4, "ACCIONISTAS", accionistas_list, [
        ("Nombre", "nombreAccionista"), ("Paterno", "apellidoPaterno"), ("Materno", "apellidoMaterno"),
        ("RFC", "rfc"), ("CURP", "curp"), ("Porcentaje", "porcentaje"), ("Dirección", "direccion1")
    ])

    # ---------------------------------------------------------
    # 6. LÍNEAS DE CRÉDITO PROCESADAS
    # ---------------------------------------------------------
    ws4.append(["LÍNEAS DE CRÉDITO ACTIVAS E HISTÓRICAS"])
    ws4.merge_cells(start_row=ws4.max_row, start_column=1, end_row=ws4.max_row, end_column=9)
    estilo_header(ws4, ws4.max_row, 1, 9)
    
    headers_lines = [
        "Institución", "Tipo Cuenta", "Límite Crédito", "Saldo Actual", 
        "Saldo Vencido", "Frecuencia Pago", "Fecha Apertura", "Último Pago", "Histórico Pagos"
    ]
    ws4.append(headers_lines)
    sub_head_row = ws4.max_row
    for col in range(1, 10):
        cell = ws4.cell(row=sub_head_row, column=col)
        cell.font = header_font
        cell.fill = sub_header_fill
        cell.alignment = Alignment(horizontal='center')

    lines = buro_info.get("credit_lines", [])
    if not lines:
        ws4.append(["No se encontró información de créditos."])
    else:
        start_data_row = ws4.max_row + 1
        for line in lines:
            ws4.append([
                line.get("institution"),
                line.get("account_type"),
                line.get("credit_limit"),
                line.get("current_balance"),
                line.get("past_due_balance"),
                line.get("payment_frequency"),
                line.get("opening_date"),
                line.get("last_payment_date"),
                line.get("payment_history")
            ])
            # Alerta visual si hay saldo vencido > 0
            if line.get("past_due_balance", 0) > 0:
                ws4.cell(row=ws4.max_row, column=5).font = Font(color="FF0000", bold=True)
                
        # Formato moneda
        for row in ws4.iter_rows(min_row=start_data_row, min_col=3, max_col=5):
            for cell in row: cell.style = currency_style

    ws4.append([""])

    # ---------------------------------------------------------
    # 7. HISTORIAL DE CONSULTAS PROCESADAS
    # ---------------------------------------------------------
    ws4.append(["HISTORIAL DE CONSULTAS (CREDIT PULLS)"])
    ws4.merge_cells(start_row=ws4.max_row, start_column=1, end_row=ws4.max_row, end_column=4)
    estilo_header(ws4, ws4.max_row, 1, 4)
    
    headers_inq = ["Institución Solicitante", "Fecha Consulta", "Tipo Contrato", "Importe Solicitado"]
    ws4.append(headers_inq)
    sub_head_row = ws4.max_row
    for col in range(1, 5):
        cell = ws4.cell(row=sub_head_row, column=col)
        cell.font = header_font
        cell.fill = sub_header_fill
        cell.alignment = Alignment(horizontal='center')

    inquiries = buro_info.get("inquiries", [])
    if not inquiries:
        ws4.append(["No hay consultas recientes registradas."])
    else:
        start_data_row = ws4.max_row + 1
        for inq in inquiries:
            ws4.append([
                inq.get("institution"),
                inq.get("inquiry_date"),
                inq.get("contract_type"),
                inq.get("amount")
            ])
        # Formato moneda
        for row in ws4.iter_rows(min_row=start_data_row, min_col=4, max_col=4):
            for cell in row: cell.style = currency_style

    # Ajuste dinámico de anchos
    ws4.column_dimensions['A'].width = 35
    ws4.column_dimensions['B'].width = 20
    ws4.column_dimensions['C'].width = 25
    ws4.column_dimensions['D'].width = 20
    ws4.column_dimensions['E'].width = 15
    ws4.column_dimensions['F'].width = 15
    ws4.column_dimensions['G'].width = 15
    ws4.column_dimensions['H'].width = 15
    ws4.column_dimensions['I'].width = 20

    # ==========================================
    # GUARDAR
    # ==========================================
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()