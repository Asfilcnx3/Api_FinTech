# utils/xlsx_converter.py

import io, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle
from typing import Dict, Any
from .helpers_texto_fluxo import (
    AGREGADORES_MAPPING,
    TERMINALES_BANCO_MAPPING
)

def generar_excel_reporte(data_json: Dict[str, Any]) -> bytes:
    wb = Workbook()
    
    # --- ESTILOS ---
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    currency_style = NamedStyle(name='currency_style', number_format='$#,##0.00')

    def aplicar_estilo_header(ws):
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

    # --- LÓGICA DE DETECCIÓN DE PROVEEDOR (SOLO PARA LA COLUMNA EXTRA DE TPV) ---
    def detectar_proveedor_terminal(descripcion: str, banco_actual: str) -> str:
        desc_lower = descripcion.lower()
        banco_upper = banco_actual.upper() if banco_actual else "GENERICO"

        # 1. Regex Banorte
        if banco_upper == "BANORTE":
            if re.search(r"\b\d{8}[cd]\b", desc_lower):
                return "BANORTE TERMINAL"

        # 2. Agregadores Globales
        for nombre_agg, keywords in AGREGADORES_MAPPING.items():
            if any(k in desc_lower for k in keywords):
                return nombre_agg

        # 3. Terminales del Banco
        if banco_upper in TERMINALES_BANCO_MAPPING:
            keywords_banco = TERMINALES_BANCO_MAPPING[banco_upper]
            if any(k in desc_lower for k in keywords_banco):
                return banco_upper 

        return "NO DEFINIDA"

    resultados = data_json.get("resultados_individuales", [])

    # ==========================================
    # 1. RESUMEN POR CUENTA
    # ==========================================
    ws1 = wb.active
    ws1.title = "Resumen por Cuenta"
    ws1.append([
        "Mes", "Cuenta", "Moneda", "Depósitos", "Cargos", "TPV Bruto", 
        "Financiamientos", "Efectivo", "Traspaso entre cuentas", "BMR CASH", "Moratorios"
    ])
    
    for res in resultados:
        ia = res.get("AnalisisIA") or {}
        if not ia: continue
        
        periodo = ia.get("periodo_fin") or ia.get("periodo_inicio") or "Desc."
        banco = ia.get("banco", "BANCO")
        clabe = str(ia.get("clabe_interbancaria") or "")
        cuenta_str = f"{banco}-{clabe[-4:]}" if len(clabe) >= 4 else banco
        
        ws1.append([
            periodo, cuenta_str,
            ia.get("tipo_moneda", "MXN"),
            ia.get("depositos", 0.0),
            ia.get("cargos", 0.0),
            ia.get("entradas_TPV_bruto", 0.0),
            ia.get("total_entradas_financiamiento", 0.0),
            ia.get("depositos_en_efectivo", 0.0),
            ia.get("traspaso_entre_cuentas", 0.0),
            ia.get("entradas_bmrcash", 0.0),
            ia.get("total_moratorios", 0.0)
        ])

    aplicar_estilo_header(ws1)
    ws1.column_dimensions['B'].width = 25
    for row in ws1.iter_rows(min_row=2, min_col=4, max_col=11):
        for cell in row: cell.style = currency_style

    # ==========================================
    # 2. RESUMEN PORTADAS
    # ==========================================
    ws2 = wb.create_sheet("Resumen Portadas")
    ws2.append([
        "Banco", "RFC", "Cliente", "CLABE / Cuenta", 
        "Periodo Inicio", "Periodo Fin", 
        "Depósitos", "Cargos", "Saldo Promedio", "Comisiones"
    ])
    
    for res in resultados:
        ia = res.get("AnalisisIA") or {}
        clabe_segura = str(ia.get("clabe_interbancaria") or "")
        ws2.append([
            ia.get("banco", "Desconocido"), ia.get("rfc", ""), ia.get("nombre_cliente", ""),
            clabe_segura, ia.get("periodo_inicio", ""), ia.get("periodo_fin", ""),
            ia.get("depositos", 0.0), ia.get("cargos", 0.0),
            ia.get("saldo_promedio", 0.0), ia.get("comisiones", 0.0)
        ])
    
    aplicar_estilo_header(ws2)
    for row in ws2.iter_rows(min_row=2, min_col=7, max_col=10):
        for cell in row: cell.style = currency_style

    # ==========================================
    # HELPER GENERADOR DE HOJAS SIMPLIFICADO
    # ==========================================
    def crear_hoja_detalle(nombre_hoja, categoria_filtro):
        """
        Crea una hoja filtrando DIRECTAMENTE por la etiqueta 'categoria' 
        que viene del JSON. Ya no recalcula reglas.
        """
        ws = wb.create_sheet(nombre_hoja)
        
        headers = ["Banco", "Fecha", "Descripción", "Monto", "Tipo", "Categoría"]
        es_hoja_tpv = (nombre_hoja == "Transacciones TPV")
        
        if es_hoja_tpv:
            headers.append("Terminal / Proveedor")
            
        ws.append(headers)
        
        for res in resultados:
            ia = res.get("AnalisisIA") or {}
            banco_actual_doc = ia.get("banco", "Desconocido")
            detalle = res.get("DetalleTransacciones", {})
            transacciones = detalle.get("transacciones", [])
            
            if isinstance(transacciones, list):
                for tx in transacciones:
                    # Datos básicos
                    cat_tx = str(tx.get("categoria", "GENERAL")).upper()
                    
                    # FILTRO MAESTRO: ¿Coincide la categoría?
                    # Si categoria_filtro es None, pasa todo (Hoja "Todos los Movimientos")
                    if categoria_filtro is not None and cat_tx != categoria_filtro:
                        continue

                    # Preparar fila
                    try:
                        monto_val = float(str(tx.get("monto", "0")).replace(",", ""))
                    except: monto_val = 0.0

                    fila = [
                        banco_actual_doc, 
                        tx.get("fecha", ""), 
                        tx.get("descripcion", ""),
                        monto_val, 
                        tx.get("tipo", ""), 
                        cat_tx
                    ]

                    # Columna extra para TPV
                    if es_hoja_tpv:
                        nombre_terminal = detectar_proveedor_terminal(tx.get("descripcion", ""), banco_actual_doc)
                        fila.append(nombre_terminal)

                    ws.append(fila)
        
        aplicar_estilo_header(ws)
        ws.column_dimensions['C'].width = 60
        if es_hoja_tpv: ws.column_dimensions['G'].width = 25
        for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
            for cell in row: cell.style = currency_style

    # ==========================================
    # DEFINICIÓN DE HOJAS (Ahora es trivial)
    # ==========================================

    # 3. TODOS LOS MOVIMIENTOS (Filtro None = Todo)
    crear_hoja_detalle("Todos los Movimientos", None)

    # 4. TRANSACCIONES TPV
    crear_hoja_detalle("Transacciones TPV", "TPV")

    # 5. EFECTIVO
    crear_hoja_detalle("Efectivo", "EFECTIVO")

    # 6. FINANCIAMIENTOS
    crear_hoja_detalle("Financiamientos", "FINANCIAMIENTO")

    # 7. TRASPASO ENTRE CUENTAS
    crear_hoja_detalle("Traspaso entre Cuentas", "TRASPASO")

    # 8. BMRCASH
    crear_hoja_detalle("BMRCASH", "BMRCASH")

    # 9. MORATORIOS
    crear_hoja_detalle("Moratorios", "MORATORIOS")

    # ==========================================
    # 10. RESUMEN GENERAL
    # ==========================================
    ws_final = wb.create_sheet("Resumen General")
    ws_final.append(["Métrica", "Valor"])
    
    total_dep = data_json.get("total_depositos", 0.0)
    es_mayor = "SÍ" if data_json.get("es_mayor_a_250") else "NO"
    
    ws_final.append(["Total Depósitos Calculados", total_dep])
    ws_final.append(["¿Es Mayor a 250k?", es_mayor])
    ws_final.append(["Documentos Procesados", len(resultados)])
    
    ws_final['B2'].style = currency_style
    aplicar_estilo_header(ws_final)
    ws_final.column_dimensions['A'].width = 30

    # ==========================================
    # 11. REPORTE TÉCNICO (QA)
    # ==========================================
    ws_qa = wb.create_sheet("Métricas de Extracción (QA)")
    
    # Encabezados Técnicos
    headers_qa = [
        "Archivo", "Cuenta", "Página", 
        "Estado", "Tiempo (s)", "Confianza (%)", 
        "Transacciones", "Bloques Texto", "Alertas del Motor"
    ]
    ws_qa.append(headers_qa)
    
    # Estilos condicionales simples
    fill_ok = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Verde
    fill_warn = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") # Amarillo
    fill_error = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Rojo

    for res in resultados:
        # Datos generales del archivo
        ia = res.get("AnalisisIA") or {}
        nombre_archivo = ia.get("nombre_archivo_virtual", "Desconocido")
        banco = ia.get("banco", "N/A")
        
        # Extraemos la metadata técnica que inyectamos en el orquestador
        # Nota: Asegúrate de que tu modelo Pydantic o Dict tenga este campo
        metadata_list = res.get("metadata_tecnica", []) 
        
        if not metadata_list:
            # Si es un archivo OCR antiguo o falló, ponemos una línea genérica
            ws_qa.append([nombre_archivo, banco, "N/A", "Sin Métricas (Legacy/OCR)", 0, 0, 0, 0, "-"])
            continue
            
        for meta in metadata_list:
            score = meta.get("calidad_score", 0.0)
            tiempo_sec = meta.get("tiempo_ms", 0) / 1000.0
            alertas = meta.get("alertas", "OK")
            
            row_data = [
                nombre_archivo,
                banco,
                meta.get("pagina", 0),
                "PROCESADO",
                tiempo_sec,
                score, # Se formateará como porcentaje
                meta.get("transacciones", 0),
                meta.get("bloques", 0),
                alertas
            ]
            ws_qa.append(row_data)
            
            # Coloreado semántico de la última fila agregada
            last_row = ws_qa[ws_qa.max_row]
            
            # Color basado en score
            c_score = last_row[5] # Columna F (Confianza)
            if score >= 0.9: c_score.fill = fill_ok
            elif score >= 0.7: c_score.fill = fill_warn
            else: c_score.fill = fill_error
            
            # Color basado en alertas
            c_alert = last_row[8] # Columna I (Alertas)
            if alertas != "OK": c_alert.fill = fill_warn

    # Formateo Final QA
    aplicar_estilo_header(ws_qa)
    ws_qa.column_dimensions['A'].width = 35
    ws_qa.column_dimensions['I'].width = 50
    
    # Formato de porcentaje para la columna de confianza
    for row in ws_qa.iter_rows(min_row=2, min_col=6, max_col=6):
        for cell in row: 
            cell.number_format = '0%'

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()